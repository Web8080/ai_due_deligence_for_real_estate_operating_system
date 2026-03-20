# Author: Victor.I
from pathlib import Path
from typing import Dict, List, Optional
from datetime import UTC, datetime
import csv
import io
import json
import os
import secrets

from fastapi import Depends, FastAPI, File, HTTPException, Header, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

from . import models, schemas
from .crm_email_import import parse_investor_email_text
from .auth import (
    build_entra_authorize_url,
    consume_auth_grant,
    consume_entra_state,
    create_default_admin,
    create_user_session,
    entra_user_allowed,
    exchange_entra_code,
    get_auth_providers,
    hash_password,
    issue_auth_grant,
    local_recovery_enabled,
    local_signup_enabled,
    map_entra_role,
    microsoft_auth_configured,
    product_demo_mode,
    require_auth,
    require_roles,
    revoke_session,
    validate_entra_id_token,
    verify_password,
)
from .database import Base, SessionLocal, engine, ensure_sqlite_compat_schema, get_db
from .ocr import extract_text
from .rag import (
    chunk_text,
    current_ai_model,
    current_ai_provider,
    embed_text,
    generate_grounded_answer,
    generate_workspace_answer,
    ollama_server_reachable,
    retrieve_top_chunks,
    store_chunks,
    _local_embed,
)

Base.metadata.create_all(bind=engine)
ensure_sqlite_compat_schema()
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

app = FastAPI(title="REOS API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:3001",
        "http://127.0.0.1:3001",
        "http://localhost:30001",
        "http://127.0.0.1:30001",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _get_or_create_company(
    db: Session,
    *,
    company_name: str | None,
    company_type: str = "investor",
    investor_type: str | None = None,
) -> models.Company | None:
    if not company_name:
        return None
    company = db.query(models.Company).filter(models.Company.name == company_name.strip()).first()
    if company:
        if investor_type and not company.investor_type:
            company.investor_type = investor_type
        return company
    company = models.Company(
        name=company_name.strip(),
        company_type=company_type,
        investor_type=investor_type,
    )
    db.add(company)
    db.flush()
    return company


def _build_analytics_summary(db: Session) -> schemas.AnalyticsSummary:
    total_deals = db.query(models.Deal).count()
    total_contacts = db.query(models.Contact).count()
    total_documents = db.query(models.Document).count()
    rows = db.query(models.Deal.stage, func.count(models.Deal.id)).group_by(models.Deal.stage).all()
    stage_distribution = {stage: count for stage, count in rows}
    return schemas.AnalyticsSummary(
        total_deals=total_deals,
        total_contacts=total_contacts,
        total_documents=total_documents,
        stage_distribution=stage_distribution,
    )


def _operations_summary(db: Session) -> schemas.WorkspaceOperationsSummary:
    high_priority_items: List[str] = []
    overdue_like_items: List[str] = []
    high_priority_deals = (
        db.query(models.Deal)
        .filter(models.Deal.priority == "high")
        .order_by(models.Deal.id.desc())
        .limit(5)
        .all()
    )
    for deal in high_priority_deals:
        high_priority_items.append(
            f"Deal #{deal.id} {deal.name} is marked high priority and next action is {deal.next_action or 'not set'}."
        )
    open_diligence = (
        db.query(models.DiligenceItem)
        .filter(models.DiligenceItem.status != "done")
        .order_by(models.DiligenceItem.id.desc())
        .limit(5)
        .all()
    )
    for item in open_diligence:
        overdue_like_items.append(
            f"Diligence item '{item.title}' for deal #{item.deal_id} remains {item.status} with severity {item.severity}."
        )
    return schemas.WorkspaceOperationsSummary(
        high_priority_items=high_priority_items,
        overdue_like_items=overdue_like_items,
    )


def _pipeline_status_from_email_decision(hint: str) -> str | None:
    h = (hint or "unknown").lower().strip()
    if h == "committed":
        return "committed"
    if h == "pass":
        return "passed"
    if h == "interested":
        return "interested"
    if h == "follow_up":
        return "contacted"
    return None


def _contact_to_crm_row(contact: models.Contact) -> schemas.CRMContactRow:
    return schemas.CRMContactRow(
        id=contact.id,
        full_name=contact.full_name,
        email=contact.email,
        contact_type=contact.contact_type,
        title=contact.title,
        phone=contact.phone,
        investor_type=contact.investor_type,
        company_id=contact.company_id,
        deal_id=contact.deal_id,
        notes=contact.notes,
        created_at=contact.created_at,
        company_name=contact.company.name if contact.company else None,
    )


def _record_audit_event(
    db: Session,
    *,
    actor: str,
    action: str,
    entity_type: str,
    entity_id: str | None = None,
    detail: str | None = None,
) -> None:
    db.add(
        models.AuditEvent(
            actor=actor,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            detail=detail,
        )
    )


def _ensure_workflow_records(db: Session) -> None:
    if db.query(models.WorkflowTask).count() > 0 or db.query(models.WorkflowException).count() > 0:
        return
    active_deals = (
        db.query(models.Deal)
        .filter(models.Deal.stage.in_(["Lead", "Screening", "Due Diligence", "Investment Committee", "Approved"]))
        .order_by(models.Deal.id.asc())
        .limit(10)
        .all()
    )
    for deal in active_deals:
        db.add(
            models.WorkflowTask(
                deal_id=deal.id,
                title=f"Advance {deal.name} through {deal.stage}",
                status="open",
                priority=deal.priority or "medium",
                owner_name=deal.owner_name,
                workflow_stage=deal.stage,
                due_context=deal.next_action or "Assign next action in workflow lane.",
            )
        )
        if deal.priority == "high" or deal.stage in {"Due Diligence", "Investment Committee"}:
            db.add(
                models.WorkflowException(
                    deal_id=deal.id,
                    title=f"{deal.name} requires operating review",
                    severity="high" if deal.priority == "high" else "medium",
                    status="open",
                    summary=f"{deal.stage} has outstanding diligence or decision pressure.",
                )
            )
    db.commit()


def _portfolio_context(db: Session) -> str:
    deals = db.query(models.Deal).order_by(models.Deal.id.desc()).limit(10).all()
    contacts = db.query(models.Contact).order_by(models.Contact.id.desc()).limit(10).all()
    investor_pipeline = db.query(models.InvestorPipelineEntry).order_by(models.InvestorPipelineEntry.id.desc()).limit(15).all()
    lines = [
        "Portfolio context:",
        *(f"Deal {deal.name} in {deal.city or 'Unknown'} is at {deal.stage} with priority {deal.priority}." for deal in deals),
        *(f"Contact {contact.full_name} supports {contact.contact_type} coverage." for contact in contacts[:5]),
        *(
            f"Investor pipeline entry {entry.status} with conviction {entry.conviction or 'unknown'} and amount {entry.commitment_amount or 0}."
            for entry in investor_pipeline[:8]
        ),
    ]
    return "\n".join(lines)


async def _ai_briefing(db: Session, workspace: str, actor: str, context_text: str, prompt: str) -> str:
    answer = await generate_workspace_answer(prompt=prompt, workspace=workspace, context=context_text)
    db.add(
        models.AIRun(
            workspace=workspace,
            actor=actor,
            provider=current_ai_provider(),
            model_name=current_ai_model(),
            prompt=prompt,
            answer=answer,
            citations="",
            status="completed",
        )
    )
    db.commit()
    return answer


def _read_only_briefing(lines: list[str], fallback: str) -> str:
    selected = [line for line in lines if line][:3]
    if not selected:
        return fallback
    return " ".join(selected)


# Named contacts for CRM / IR demos (paired with example.com emails).
_CONTACT_IDENTITY_ROWS: List[tuple[str, str, str]] = [
    ("Jordan Ellis", "principal", "industrial"),
    ("Priya Shah", "director", "medical office"),
    ("Marcus Webb", "head of acquisitions", "multifamily"),
    ("Elena Varga", "managing director", "debt"),
    ("Sam Rivera", "analyst", "multifamily"),
    ("Taylor Brooks", "associate", "office"),
    ("Riley Morgan", "vp", "mixed use"),
    ("Casey Nguyen", "associate", "logistics"),
    ("Morgan Patel", "director ir", "institutional"),
    ("Alex Chen", "partner", "family office"),
    ("Jamie Ortiz", "principal", "fund of funds"),
    ("Drew Kim", "counsel", "legal"),
    ("Blake Foster", "md", "lending"),
    ("Quinn Murphy", "vp", "retail"),
    ("Reese Park", "analyst", "hospitality"),
    ("Avery Cole", "principal", "self storage"),
    ("Skyler Diaz", "associate", "life sciences"),
    ("Emerson Gray", "director", "senior housing"),
    ("Finley Reed", "vp", "industrial"),
    ("Harper Lane", "analyst", "core office"),
    ("Rowan West", "principal", "value add"),
    ("Sage Moore", "director", "capital markets"),
    ("River James", "associate", "transactions"),
    ("Phoenix Hill", "partner", "broker coverage"),
]


def _backfill_demo_dataset(db: Session) -> Dict[str, int]:
    """Idempotent enrichment so governance, deal workspace, RAG, and operations pages stay populated."""
    counts = {
        "workflow_tasks_added": 0,
        "workflow_exceptions_added": 0,
        "deal_notes_added": 0,
        "documents_added": 0,
        "ai_query_logs_added": 0,
        "ai_runs_added": 0,
        "audit_events_added": 0,
        "chunks_added": 0,
    }
    deals = db.query(models.Deal).order_by(models.Deal.id.asc()).all()
    if not deals:
        return counts

    active_stages = frozenset({"Lead", "Screening", "Due Diligence", "Investment Committee", "Approved", "Closing"})
    active_deals = [d for d in deals if d.stage in active_stages]

    task_deal_ids = {
        row[0]
        for row in db.query(models.WorkflowTask.deal_id)
        .filter(models.WorkflowTask.deal_id.isnot(None))
        .distinct()
        .all()
    }
    target_tasks = 18
    cur_tasks = db.query(models.WorkflowTask).count()
    for deal in active_deals:
        if cur_tasks >= target_tasks:
            break
        if deal.id in task_deal_ids:
            continue
        db.add(
            models.WorkflowTask(
                deal_id=deal.id,
                title=f"Advance {deal.name} through {deal.stage}",
                status="open",
                priority=deal.priority or "medium",
                owner_name=deal.owner_name,
                workflow_stage=deal.stage,
                due_context=deal.next_action or "Assign next action in workflow lane.",
            )
        )
        task_deal_ids.add(deal.id)
        cur_tasks += 1
        counts["workflow_tasks_added"] += 1

    exc_deal_ids = {
        row[0]
        for row in db.query(models.WorkflowException.deal_id)
        .filter(models.WorkflowException.deal_id.isnot(None))
        .distinct()
        .all()
    }
    target_exc = 10
    cur_exc = db.query(models.WorkflowException).count()
    for deal in active_deals:
        if cur_exc >= target_exc:
            break
        if deal.id in exc_deal_ids:
            continue
        if deal.priority != "high" and deal.stage not in {"Due Diligence", "Investment Committee", "Closing"}:
            continue
        db.add(
            models.WorkflowException(
                deal_id=deal.id,
                title=f"{deal.name} requires operating review",
                severity="high" if deal.priority == "high" else "medium",
                status="open",
                summary=f"{deal.stage} has outstanding diligence or decision pressure.",
            )
        )
        exc_deal_ids.add(deal.id)
        cur_exc += 1
        counts["workflow_exceptions_added"] += 1

    for deal in deals[:14]:
        note_n = db.query(models.DealNote).filter(models.DealNote.deal_id == deal.id).count()
        if note_n >= 2:
            continue
        db.add(
            models.DealNote(
                deal_id=deal.id,
                author="manager1",
                content=(
                    f"Committee prep: confirm rent roll tie-out and capex schedule for {deal.name}. "
                    "Flag any sponsor deviation from OM base case."
                ),
            )
        )
        counts["deal_notes_added"] += 1
        if note_n == 0:
            db.add(
                models.DealNote(
                    deal_id=deal.id,
                    author="analyst1",
                    content=f"Site tour notes logged; broker expects IOI timeline after environmental screen.",
                )
            )
            counts["deal_notes_added"] += 1

    extra_doc_specs = [
        ("lease_abstract", "Key lease terms extracted; watch rollover in years 3-5.", "rollover,ti"),
        ("environmental_screen", "Phase I no recognized conditions; confirm no vapor intrusion pathway.", "environmental"),
    ]
    for deal in deals[:10]:
        existing = db.query(models.Document).filter(models.Document.deal_id == deal.id).count()
        if existing >= 3:
            continue
        for doc_type, summary, tags in extra_doc_specs:
            if db.query(models.Document).filter(models.Document.deal_id == deal.id, models.Document.document_type == doc_type).first():
                continue
            slug = deal.name.lower().replace(" ", "_")[:40]
            db.add(
                models.Document(
                    deal_id=deal.id,
                    filename=f"{slug}_{doc_type}.txt",
                    document_type=doc_type,
                    status="processed",
                    summary=summary,
                    risk_tags=tags,
                    content=(
                        f"{deal.name} — {doc_type.replace('_', ' ')}. "
                        f"Market: {deal.city or 'N/A'}, {deal.state or ''}. "
                        "Supporting narrative for diligence workspace and retrieval demos."
                    ),
                )
            )
            counts["documents_added"] += 1
            if db.query(models.Document).filter(models.Document.deal_id == deal.id).count() >= 3:
                break

    sample_qa = [
        ("What are the top lease risks mentioned in the materials?", "Lease rollover clusters in years 3-5; sponsor proposes partial capex reserve."),
        ("Summarize environmental posture for this deal.", "Phase I indicates no RECs; recommend confirmatory soil sampling if redevelopment planned."),
        ("What should IC focus on this week?", "Rent roll reconciliation vs OM and debt sizing sensitivity on exit cap."),
    ]
    if db.query(models.AIQueryLog).count() < 8:
        for i, deal in enumerate(deals[:4]):
            q, a = sample_qa[i % len(sample_qa)]
            db.add(
                models.AIQueryLog(
                    deal_id=deal.id,
                    username="analyst1",
                    question=q,
                    answer=a,
                    citations=json.dumps([f"deal:{deal.id}:seed"]),
                )
            )
            counts["ai_query_logs_added"] += 1

    copilot_samples = [
        ("leads", "Which screening deals need broker follow-up this week?", "Prioritize industrial and medical office where inbound is thin; pair with Apollo targets."),
        ("portfolio", "What is crowding the IC agenda?", "High-priority names in Due Diligence and IC with open environmental or title items."),
        ("crm", "Who should IR sequence next?", "Family offices with recent data room access but no commitment letter."),
        ("investors", "Where is momentum stalling?", "Contacts in contacted status without a scheduled second call in 10 days."),
    ]
    if db.query(models.AIRun).count() < 6:
        for workspace, prompt, answer in copilot_samples:
            if db.query(models.AIRun).filter(models.AIRun.workspace == workspace, models.AIRun.prompt == prompt).first():
                continue
            db.add(
                models.AIRun(
                    workspace=workspace,
                    deal_id=None,
                    actor="analyst1",
                    provider=current_ai_provider(),
                    model_name=current_ai_model(),
                    prompt=prompt,
                    answer=answer,
                    citations="",
                    status="completed",
                )
            )
            counts["ai_runs_added"] += 1

    if db.query(models.AuditEvent).filter(models.AuditEvent.action == "demo_dataset_enriched").count() < 1:
        db.add(
            models.AuditEvent(
                actor="system",
                action="demo_dataset_enriched",
                entity_type="workspace",
                entity_id="local",
                detail="Applied idempotent demo backfill for workflow, notes, documents, AI logs, and audit samples.",
            )
        )
        counts["audit_events_added"] += 1
    audit_extras = [
        ("admin", "config_view", "integration", None, "Reviewed integration catalog toggles (demo)."),
        ("manager1", "deal_stage_change", "deal", "1", "Recorded stage transition for committee tracking."),
        ("analyst1", "document_upload", "document", None, "Indexed supplemental diligence file."),
        ("admin", "login_success", "session", None, "Local session established."),
        ("manager1", "report_export", "reports", None, "Queued committee summary export."),
        ("analyst1", "diligence_update", "diligence", None, "Updated severity on open environmental item."),
    ]
    for actor, action, etype, eid, detail in audit_extras:
        if db.query(models.AuditEvent).count() >= 14:
            break
        exists = (
            db.query(models.AuditEvent)
            .filter(
                models.AuditEvent.actor == actor,
                models.AuditEvent.action == action,
                models.AuditEvent.detail == detail,
            )
            .first()
        )
        if exists:
            continue
        db.add(
            models.AuditEvent(actor=actor, action=action, entity_type=etype, entity_id=eid, detail=detail)
        )
        counts["audit_events_added"] += 1

    for deal in deals[:5]:
        doc = (
            db.query(models.Document)
            .filter(models.Document.deal_id == deal.id)
            .order_by(models.Document.id.asc())
            .first()
        )
        if not doc:
            continue
        if db.query(models.Chunk).filter(models.Chunk.document_id == doc.id).count() > 0:
            continue
        text_src = (doc.content or doc.summary or "demo").strip()
        pieces = chunk_text(text_src, size=120, overlap=20)[:3] or [text_src[:400]]
        for piece in pieces:
            emb = _local_embed(piece)
            db.add(
                models.Chunk(
                    deal_id=deal.id,
                    document_id=doc.id,
                    content=piece[:2000],
                    embedding=json.dumps(emb),
                )
            )
            counts["chunks_added"] += 1

    db.commit()
    return counts


def _seed_demo_records(db: Session) -> schemas.DemoSeedResponse:
    if db.query(models.Deal).count() > 0:
        return schemas.DemoSeedResponse(
            deals_created=db.query(models.Deal).count(),
            contacts_created=db.query(models.Contact).count(),
            investor_pipeline_entries_created=db.query(models.InvestorPipelineEntry).count(),
            workflow_tasks_added=0,
            workflow_exceptions_added=0,
            deal_notes_added=0,
            documents_added=0,
            ai_query_logs_added=0,
            ai_runs_added=0,
            audit_events_added=0,
            chunks_added=0,
        )

    stage_templates = [
        ("Maple Grove Multifamily", "Screening", "multifamily", "Dallas", "TX", "Broker referral", "high"),
        ("West Loop Office", "Due Diligence", "office", "Chicago", "IL", "Direct outreach", "high"),
        ("Harbor Logistics Park", "Investment Committee", "industrial", "Savannah", "GA", "Broker referral", "medium"),
        ("Sunset Retail Center", "Lead", "retail", "Phoenix", "AZ", "Inbound", "medium"),
        ("Riverfront Hotel", "Approved", "hospitality", "Nashville", "TN", "Broker referral", "high"),
        ("Civic Medical Plaza", "Due Diligence", "medical office", "Austin", "TX", "Seller network", "high"),
        ("Pine Distribution Hub", "Screening", "industrial", "Memphis", "TN", "Inbound", "medium"),
        ("Northshore Self Storage", "Lead", "self storage", "Tampa", "FL", "Cold outreach", "low"),
        ("Broadway Mixed Use", "Rejected", "mixed use", "New York", "NY", "Broker referral", "medium"),
        ("Lakeside Senior Living", "Screening", "senior housing", "Charlotte", "NC", "Inbound", "high"),
        ("Mission Bay Lab Space", "Due Diligence", "life sciences", "San Diego", "CA", "Direct outreach", "medium"),
        ("Capitol Hill Residential", "Investment Committee", "multifamily", "Seattle", "WA", "Broker referral", "high"),
        ("Bayview Logistics II", "Closing", "industrial", "Jacksonville", "FL", "Broker referral", "high"),
        ("Riverside Medical MOB", "Screening", "medical office", "Riverside", "CA", "Inbound", "medium"),
        ("Highland Data Center", "Lead", "industrial", "Ashburn", "VA", "Direct outreach", "high"),
    ]
    deal_records = []
    for idx, (name, stage, asset_type, city, state, source, priority) in enumerate(stage_templates, start=1):
        deal = models.Deal(
            name=name,
            stage=stage,
            description=f"{asset_type.title()} opportunity in {city} requiring structured diligence and investor coordination.",
            asset_type=asset_type,
            city=city,
            state=state,
            source=source,
            priority=priority,
            owner_name="analyst1" if idx % 2 else "manager1",
            next_action="Review diligence pack" if stage != "Approved" else "Prepare closing handoff",
        )
        db.add(deal)
        db.flush()
        db.add(
            models.DealStageEvent(
                deal_id=deal.id,
                from_stage=None,
                to_stage=stage,
                reason="Seeded local MVP workflow state",
                author="system",
            )
        )
        for item_idx in range(3):
            st = "done" if item_idx == 2 else ("open" if item_idx == 0 else "in_review")
            db.add(
                models.DiligenceItem(
                    deal_id=deal.id,
                    title=f"{name} diligence item {item_idx + 1}",
                    status=st,
                    severity="high" if item_idx == 0 and priority == "high" else "medium",
                    owner_name="analyst1" if item_idx % 2 == 0 else "manager1",
                    notes="Follow up with seller or legal counsel for missing support.",
                )
            )
        if idx <= 4 and stage not in {"Lead"}:
            db.add(
                models.DealStageEvent(
                    deal_id=deal.id,
                    from_stage="Lead",
                    to_stage=stage,
                    reason="Seeded progression snapshot",
                    author="system",
                )
            )
        db.add(
            models.Document(
                deal_id=deal.id,
                filename=f"{name.lower().replace(' ', '_')}_summary.txt",
                document_type="offering_memo",
                status="processed",
                summary=f"{name} summary generated for local MVP demo.",
                risk_tags="lease rollover,capex",
                content=f"{name} contains lease rollover exposure and capital planning items that should be reviewed.",
            )
        )
        deal_records.append(deal)

    company_templates = [
        ("Northbridge Capital", "investor", "family office"),
        ("Elm Street Partners", "investor", "private equity"),
        ("Harborview Advisors", "investor", "fund of funds"),
        ("Summit Brokerage", "broker", None),
        ("Atlas Legal", "legal", None),
        ("Stonebank Lending", "lender", None),
        ("Blue Pine Capital", "investor", "institutional"),
        ("Crescent Family Office", "investor", "family office"),
    ]
    companies = {}
    for name, company_type, investor_type in company_templates:
        company = _get_or_create_company(
            db,
            company_name=name,
            company_type=company_type,
            investor_type=investor_type,
        )
        companies[name] = company

    contacts_created = 0
    investor_entries_created = 0
    statuses = ["target", "contacted", "interested", "passed", "committed"]
    for idx in range(24):
        company_name = list(companies.keys())[idx % len(companies)]
        company = companies[company_name]
        id_row = _CONTACT_IDENTITY_ROWS[idx % len(_CONTACT_IDENTITY_ROWS)]
        display_name, title_key, _mandate = id_row
        title_display = title_key.replace("_", " ").title()
        contact = models.Contact(
            full_name=display_name,
            email=f"{display_name.lower().replace(' ', '.')}@example.com",
            contact_type="investor" if company.company_type == "investor" else company.company_type,
            title=title_display if company.company_type == "investor" else f"{title_display} ({company.company_type})",
            phone=f"+1-555-010-{idx:02d}",
            investor_type=company.investor_type,
            company_id=company.id,
            deal_id=deal_records[idx % len(deal_records)].id,
            notes="Imported or seeded contact for local MVP review.",
        )
        db.add(contact)
        db.flush()
        contacts_created += 1
        if company.company_type == "investor":
            db.add(
                models.InvestorPipelineEntry(
                    deal_id=deal_records[idx % len(deal_records)].id,
                    contact_id=contact.id,
                    status=statuses[idx % len(statuses)],
                    commitment_amount=250000 * ((idx % 4) + 1),
                    conviction="high" if idx % 3 == 0 else "medium",
                    last_signal="Opened teaser and requested follow-up" if idx % 2 == 0 else "Awaiting reply",
                    next_action="Schedule follow-up call" if idx % 2 == 0 else "Send diligence memo",
                )
            )
            investor_entries_created += 1

    for comp in list(companies.values())[:4]:
        if not comp.notes:
            comp.notes = "Seeded relationship; use CRM graph and diligence links for coverage."

    db.commit()
    return schemas.DemoSeedResponse(
        deals_created=len(deal_records),
        contacts_created=contacts_created,
        investor_pipeline_entries_created=investor_entries_created,
        workflow_tasks_added=0,
        workflow_exceptions_added=0,
        deal_notes_added=0,
        documents_added=0,
        ai_query_logs_added=0,
        ai_runs_added=0,
        audit_events_added=0,
        chunks_added=0,
    )


def _integration_status_snapshot() -> schemas.IntegrationStatus:
    runtime_mode = os.getenv("REOS_RUNTIME_MODE", "local").strip().lower()
    azure_blob_configured = bool(os.getenv("REOS_AZURE_STORAGE_ACCOUNT") and os.getenv("REOS_AZURE_STORAGE_CONTAINER"))
    azure_ad_configured = microsoft_auth_configured() or bool(
        os.getenv("REOS_AZURE_TENANT_ID") and os.getenv("REOS_AZURE_CLIENT_ID") and os.getenv("REOS_AZURE_AUDIENCE")
    )
    azure_key_vault_configured = bool(os.getenv("REOS_AZURE_KEY_VAULT_URL"))
    azure_front_door_configured = bool(os.getenv("REOS_AZURE_FRONT_DOOR_HOST"))
    azure_app_gateway_configured = bool(os.getenv("REOS_AZURE_APP_GATEWAY_HOST"))
    azure_api_management_configured = bool(os.getenv("REOS_AZURE_APIM_NAME"))
    azure_service_bus_configured = bool(
        os.getenv("REOS_AZURE_SERVICE_BUS_NAMESPACE") and os.getenv("REOS_AZURE_SERVICE_BUS_QUEUE")
    )
    azure_ai_search_configured = bool(
        os.getenv("REOS_AZURE_AI_SEARCH_ENDPOINT") and os.getenv("REOS_AZURE_AI_SEARCH_INDEX")
    )
    azure_functions_configured = bool(os.getenv("REOS_AZURE_FUNCTIONS_APP"))
    return schemas.IntegrationStatus(
        runtime_mode=runtime_mode,
        ai_provider=current_ai_provider(),
        azure_openai_configured=bool(
            os.getenv("REOS_AZURE_OPENAI_ENDPOINT")
            and os.getenv("REOS_AZURE_OPENAI_API_KEY")
            and os.getenv("REOS_AZURE_OPENAI_CHAT_DEPLOYMENT")
            and os.getenv("REOS_AZURE_OPENAI_EMBED_DEPLOYMENT")
        ),
        azure_blob_configured=azure_blob_configured,
        azure_ad_configured=azure_ad_configured,
        azure_key_vault_configured=azure_key_vault_configured,
        azure_front_door_configured=azure_front_door_configured,
        azure_app_gateway_configured=azure_app_gateway_configured,
        azure_api_management_configured=azure_api_management_configured,
        azure_service_bus_configured=azure_service_bus_configured,
        azure_ai_search_configured=azure_ai_search_configured,
        azure_functions_configured=azure_functions_configured,
        automation_mode=os.getenv("REOS_AUTOMATION_MODE", "assistive"),
    )


def _integration_preferences(db: Session) -> Dict[str, bool]:
    preferences = db.query(models.IntegrationPreference).all()
    return {item.key: bool(item.enabled) for item in preferences}


def _integration_definitions(status: schemas.IntegrationStatus) -> list[dict]:
    return [
        {
            "key": "azure_openai",
            "label": "Azure OpenAI",
            "category": "ai",
            "connected": status.azure_openai_configured,
            "mode": "live",
            "placeholder": False,
            "api_ready": True,
            "auth_type": "api_key",
            "summary": "Enterprise LLM and embeddings provider for governed AI mode.",
            "required_env_vars": [
                "REOS_AZURE_OPENAI_ENDPOINT",
                "REOS_AZURE_OPENAI_API_KEY",
                "REOS_AZURE_OPENAI_CHAT_DEPLOYMENT",
                "REOS_AZURE_OPENAI_EMBED_DEPLOYMENT",
            ],
            "config_fields": [
                {"key": "endpoint", "label": "Endpoint", "value_hint": "https://<resource>.openai.azure.com"},
                {"key": "chat_deployment", "label": "Chat deployment", "value_hint": "gpt-4o"},
                {"key": "embed_deployment", "label": "Embed deployment", "value_hint": "text-embedding-3-large"},
            ],
            "notes": ["Used when REOS runtime mode is switched to Azure.", "Pairs with AI Search for grounded retrieval."],
            "last_test_result": "Readiness is derived from env var checks.",
        },
        {
            "key": "azure_blob_storage",
            "label": "Azure Blob Storage",
            "category": "data",
            "connected": status.azure_blob_configured,
            "mode": "live",
            "placeholder": False,
            "api_ready": True,
            "auth_type": "connection_string",
            "summary": "Document storage target for enterprise document ingestion.",
            "required_env_vars": ["REOS_AZURE_STORAGE_ACCOUNT", "REOS_AZURE_STORAGE_CONTAINER"],
            "config_fields": [
                {"key": "account", "label": "Storage account", "value_hint": "reosstorageprod"},
                {"key": "container", "label": "Container", "value_hint": "deal-documents"},
            ],
            "notes": ["Currently modeled for readiness only.", "Future uploads should be dual-wired behind a storage adapter."],
            "last_test_result": "Env var presence check only.",
        },
        {
            "key": "azure_ai_search",
            "label": "Azure AI Search",
            "category": "ai",
            "connected": status.azure_ai_search_configured,
            "mode": "live",
            "placeholder": False,
            "api_ready": True,
            "auth_type": "api_key",
            "summary": "Managed vector retrieval target for enterprise search and RAG.",
            "required_env_vars": ["REOS_AZURE_AI_SEARCH_ENDPOINT", "REOS_AZURE_AI_SEARCH_INDEX"],
            "config_fields": [
                {"key": "endpoint", "label": "Search endpoint", "value_hint": "https://<search>.search.windows.net"},
                {"key": "index", "label": "Index name", "value_hint": "reos-documents"},
            ],
            "notes": ["Will replace local chunk retrieval at enterprise scale."],
            "last_test_result": "Env var presence check only.",
        },
        {
            "key": "azure_service_bus",
            "label": "Azure Service Bus",
            "category": "operations",
            "connected": status.azure_service_bus_configured,
            "mode": "live",
            "placeholder": False,
            "api_ready": True,
            "auth_type": "connection_string",
            "summary": "Queue backbone for asynchronous document and automation processing.",
            "required_env_vars": ["REOS_AZURE_SERVICE_BUS_NAMESPACE", "REOS_AZURE_SERVICE_BUS_QUEUE"],
            "config_fields": [
                {"key": "namespace", "label": "Namespace", "value_hint": "reos-ops-bus"},
                {"key": "queue", "label": "Queue", "value_hint": "document-ingestion"},
            ],
            "notes": ["Pairs with Functions for event-driven processing."],
            "last_test_result": "Env var presence check only.",
        },
        {
            "key": "microsoft_entra_id",
            "label": "Microsoft Entra ID",
            "category": "security",
            "connected": status.azure_ad_configured,
            "mode": "live",
            "placeholder": False,
            "api_ready": True,
            "auth_type": "oauth",
            "summary": "Enterprise identity provider for future SSO and role mapping.",
            "required_env_vars": [
                "REOS_ENTRA_TENANT_ID",
                "REOS_ENTRA_CLIENT_ID",
                "REOS_ENTRA_CLIENT_SECRET",
                "REOS_ENTRA_REDIRECT_URI",
                "REOS_ENTRA_FRONTEND_CALLBACK",
            ],
            "config_fields": [
                {"key": "tenant_id", "label": "Tenant ID", "value_hint": "xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx"},
                {"key": "client_id", "label": "Client ID", "value_hint": "application registration id"},
            ],
            "notes": [
                "Frontend login can redirect to Microsoft once tenant credentials are configured.",
                "Role mapping is currently driven by configured admin and manager email allowlists.",
            ],
            "last_test_result": "Live readiness is derived from Entra environment configuration.",
        },
        {
            "key": "microsoft_graph",
            "label": "Microsoft Graph",
            "category": "communications",
            "connected": False,
            "mode": "placeholder",
            "placeholder": True,
            "api_ready": False,
            "auth_type": "oauth",
            "summary": "Email, contacts, and calendar ingestion for investor and internal workflow automation.",
            "required_env_vars": ["REOS_GRAPH_CLIENT_ID", "REOS_GRAPH_CLIENT_SECRET", "REOS_GRAPH_TENANT_ID"],
            "config_fields": [
                {"key": "mailbox_scope", "label": "Mailbox scope", "value_hint": "shared mailbox or user mailbox"},
                {"key": "tenant_id", "label": "Tenant ID", "value_hint": "Azure tenant id"},
                {"key": "client_secret", "label": "Client secret", "value_hint": "stored in env or vault", "secret": True},
            ],
            "notes": [
                "Use for investor contact import, meeting sync, and email signal detection.",
                "Displayed as a placeholder until API integration is implemented.",
            ],
            "last_test_result": "Placeholder only. No live connectivity yet.",
        },
        {
            "key": "gmail",
            "label": "Gmail",
            "category": "communications",
            "connected": False,
            "mode": "placeholder",
            "placeholder": True,
            "api_ready": False,
            "auth_type": "oauth",
            "summary": "Alternative email ingestion path for firms outside Microsoft 365.",
            "required_env_vars": ["REOS_GMAIL_CLIENT_ID", "REOS_GMAIL_CLIENT_SECRET"],
            "config_fields": [
                {"key": "workspace_domain", "label": "Workspace domain", "value_hint": "firm.com"},
                {"key": "client_secret", "label": "Client secret", "value_hint": "stored in env or vault", "secret": True},
            ],
            "notes": ["Mirrors Graph use cases with Gmail and Google Calendar."],
            "last_test_result": "Placeholder only. No live connectivity yet.",
        },
        {
            "key": "slack",
            "label": "Slack",
            "category": "operations",
            "connected": False,
            "mode": "placeholder",
            "placeholder": True,
            "api_ready": False,
            "auth_type": "bot_token",
            "summary": "Alerting and workflow notifications for diligence and investor events.",
            "required_env_vars": ["REOS_SLACK_BOT_TOKEN", "REOS_SLACK_SIGNING_SECRET"],
            "config_fields": [
                {"key": "channel", "label": "Default channel", "value_hint": "#reos-ops"},
                {"key": "bot_token", "label": "Bot token", "value_hint": "xoxb-...", "secret": True},
            ],
            "notes": ["Good first notification placeholder for QA and product review."],
            "last_test_result": "Placeholder only. No live connectivity yet.",
        },
        {
            "key": "docusign",
            "label": "DocuSign",
            "category": "signatures",
            "connected": False,
            "mode": "placeholder",
            "placeholder": True,
            "api_ready": False,
            "auth_type": "oauth",
            "summary": "Signature workflow integration for closings, approvals, and subscription documents.",
            "required_env_vars": ["REOS_DOCUSIGN_CLIENT_ID", "REOS_DOCUSIGN_ACCOUNT_ID"],
            "config_fields": [
                {"key": "account_id", "label": "Account ID", "value_hint": "DocuSign account id"},
                {"key": "template_set", "label": "Template set", "value_hint": "closing-pack-v1"},
            ],
            "notes": ["Placeholder for legal and closing workflow testing."],
            "last_test_result": "Placeholder only. No live connectivity yet.",
        },
        {
            "key": "alloy",
            "label": "Alloy",
            "category": "compliance",
            "connected": False,
            "mode": "placeholder",
            "placeholder": True,
            "api_ready": False,
            "auth_type": "api_key",
            "summary": "KYC and compliance verification placeholder for investor onboarding.",
            "required_env_vars": ["REOS_ALLOY_API_KEY", "REOS_ALLOY_WORKFLOW_TOKEN"],
            "config_fields": [
                {"key": "workflow_token", "label": "Workflow token", "value_hint": "kyc-investor-v1"},
                {"key": "api_key", "label": "API key", "value_hint": "stored in env or vault", "secret": True},
            ],
            "notes": ["Useful for onboarding and compliance workflow demos before vendor setup."],
            "last_test_result": "Placeholder only. No live connectivity yet.",
        },
        {
            "key": "juniper_square",
            "label": "Juniper Square",
            "category": "investor_systems",
            "connected": False,
            "mode": "placeholder",
            "placeholder": True,
            "api_ready": False,
            "auth_type": "api_key",
            "summary": "Investor onboarding and reporting system placeholder.",
            "required_env_vars": ["REOS_JUNIPER_SQUARE_API_KEY"],
            "config_fields": [
                {"key": "workspace_id", "label": "Workspace ID", "value_hint": "fund-admin-workspace"},
                {"key": "api_key", "label": "API key", "value_hint": "stored in env or vault", "secret": True},
            ],
            "notes": ["Represents future LP onboarding and reporting integration."],
            "last_test_result": "Placeholder only. No live connectivity yet.",
        },
        {
            "key": "yardi",
            "label": "Yardi",
            "category": "market_data",
            "connected": False,
            "mode": "placeholder",
            "placeholder": True,
            "api_ready": False,
            "auth_type": "api_key",
            "summary": "Property and operational data placeholder for asset management and reporting.",
            "required_env_vars": ["REOS_YARDI_CLIENT_ID", "REOS_YARDI_CLIENT_SECRET"],
            "config_fields": [
                {"key": "property_scope", "label": "Property scope", "value_hint": "portfolio or asset ids"},
                {"key": "client_secret", "label": "Client secret", "value_hint": "stored in env or vault", "secret": True},
            ],
            "notes": ["Structured as a placeholder until downstream property systems are selected."],
            "last_test_result": "Placeholder only. No live connectivity yet.",
        },
        {
            "key": "apollo_io",
            "label": "Apollo.io",
            "category": "growth",
            "connected": bool(os.getenv("REOS_APOLLO_API_KEY", "").strip()),
            "mode": "live" if os.getenv("REOS_APOLLO_API_KEY", "").strip() else "placeholder",
            "placeholder": not bool(os.getenv("REOS_APOLLO_API_KEY", "").strip()),
            "api_ready": True,
            "auth_type": "api_key",
            "summary": "B2B contact and account data for origination prospecting and sequence-ready lists.",
            "required_env_vars": ["REOS_APOLLO_API_KEY"],
            "config_fields": [
                {"key": "api_key", "label": "API key", "value_hint": "Apollo private API key", "secret": True},
                {"key": "icp_tags", "label": "ICP tags", "value_hint": "industrial, multifamily, family-office"},
            ],
            "notes": [
                "Demo UI shows scored prospects; live sync requires a valid key and server-side rate limits.",
                "AI fit scoring runs as a governed preview until LLM routing is enabled for leads.",
            ],
            "last_test_result": "Key present" if os.getenv("REOS_APOLLO_API_KEY", "").strip() else "Not configured",
        },
        {
            "key": "snov_io",
            "label": "Snov.io",
            "category": "growth",
            "connected": bool(os.getenv("REOS_SNOV_API_KEY", "").strip()),
            "mode": "live" if os.getenv("REOS_SNOV_API_KEY", "").strip() else "placeholder",
            "placeholder": not bool(os.getenv("REOS_SNOV_API_KEY", "").strip()),
            "api_ready": True,
            "auth_type": "api_key",
            "summary": "Email discovery, verification, and drip-style outreach hooks for investor and broker lists.",
            "required_env_vars": ["REOS_SNOV_API_KEY"],
            "config_fields": [
                {"key": "api_key", "label": "API key", "value_hint": "Snov REST API key", "secret": True},
                {"key": "daily_cap", "label": "Daily find cap", "value_hint": "500"},
            ],
            "notes": [
                "Pairs with Apollo when Apollo owns accounts and Snov improves deliverability verification.",
                "Outbound sends stay behind human approval and firm compliance review.",
            ],
            "last_test_result": "Key present" if os.getenv("REOS_SNOV_API_KEY", "").strip() else "Not configured",
        },
        {
            "key": "fred_macro_rates",
            "label": "FRED (macro rates)",
            "category": "market_data",
            "connected": bool(os.getenv("REOS_FRED_API_KEY", "").strip()),
            "mode": "live" if os.getenv("REOS_FRED_API_KEY", "").strip() else "placeholder",
            "placeholder": not bool(os.getenv("REOS_FRED_API_KEY", "").strip()),
            "api_ready": True,
            "auth_type": "api_key",
            "summary": "Treasury and policy-rate series for discount-rate overlays, cap context, and IC briefing charts.",
            "required_env_vars": ["REOS_FRED_API_KEY"],
            "config_fields": [
                {"key": "api_key", "label": "API key", "value_hint": "FRED API key (St. Louis Fed)", "secret": True},
                {"key": "series_ids", "label": "Series IDs", "value_hint": "DGS10, SOFR, OBFR"},
            ],
            "notes": [
                "Baseline public macro feed before paid capital-markets terminals.",
                "Cache series aggressively; document ingestion timestamps for audit.",
            ],
            "last_test_result": "Key present" if os.getenv("REOS_FRED_API_KEY", "").strip() else "Not configured",
        },
        {
            "key": "compstak",
            "label": "CompStak",
            "category": "market_data",
            "connected": False,
            "mode": "placeholder",
            "placeholder": True,
            "api_ready": False,
            "auth_type": "api_key",
            "summary": "Comp sales and lease comps for screening and underwriting cross-checks.",
            "required_env_vars": ["REOS_COMPSTAK_API_KEY"],
            "config_fields": [
                {"key": "api_key", "label": "API key", "value_hint": "tenant-specific key", "secret": True},
                {"key": "markets", "label": "Markets", "value_hint": "NYC, Sunbelt MSAs"},
            ],
            "notes": ["Contract and data-license dependent.", "Map into comp panels on the deal workspace."],
            "last_test_result": "Placeholder only. No live connectivity yet.",
        },
        {
            "key": "institutional_market_data",
            "label": "Bloomberg / Refinitiv / ICE",
            "category": "market_data",
            "connected": False,
            "mode": "placeholder",
            "placeholder": True,
            "api_ready": False,
            "auth_type": "mixed",
            "summary": "Institutional feeds for CMBS spreads, credit indices, FX, and benchmark curves used in large-firm underwriting.",
            "required_env_vars": ["REOS_INSTITUTIONAL_MARKET_DATA_MODE"],
            "config_fields": [
                {"key": "vendor", "label": "Vendor", "value_hint": "bloomberg | refinitiv | ice"},
                {"key": "entitlement_id", "label": "Entitlement", "value_hint": "firm contract id"},
            ],
            "notes": [
                "Single catalog row for multiple enterprise vendors; implementation tracks licensed stack.",
                "Expected path: server-side adapter + entitlement checks, no credentials in the browser.",
            ],
            "last_test_result": "Placeholder only. No live connectivity yet.",
        },
    ]


def _integration_catalog_status(definition: dict, enabled: bool) -> str:
    if definition["placeholder"]:
        return "mock_enabled" if enabled else "placeholder"
    if definition["connected"] and enabled:
        return "connected"
    if definition["connected"] and not enabled:
        return "configured"
    return "missing_config"


def _build_integration_catalog(db: Session) -> schemas.IntegrationCatalogResponse:
    status = _integration_status_snapshot()
    preferences = _integration_preferences(db)
    items = []
    demo = product_demo_mode()
    demo_suffix = " Demo posture: no outbound vendor calls." if demo else ""
    for definition in _integration_definitions(status):
        default_enabled = definition["connected"] if not definition["placeholder"] else False
        enabled = preferences.get(definition["key"], default_enabled)
        last_result = definition["last_test_result"] + demo_suffix if demo else definition["last_test_result"]
        items.append(
            schemas.IntegrationCatalogItem(
                key=definition["key"],
                label=definition["label"],
                category=definition["category"],
                status=_integration_catalog_status(definition, enabled),
                enabled=enabled,
                connected=definition["connected"],
                mode=definition["mode"],
                placeholder=definition["placeholder"],
                api_ready=definition["api_ready"],
                auth_type=definition["auth_type"],
                summary=definition["summary"],
                notes=definition["notes"],
                config_fields=[schemas.IntegrationConfigField(**field) for field in definition["config_fields"]],
                required_env_vars=definition["required_env_vars"],
                last_test_result=last_result.strip(),
            )
        )
    notice = ""
    if demo:
        notice = (
            "Product demo mode is on: connectors are specifications and operator toggles only. "
            "Nothing in this catalog opens live sessions to vendors until environment credentials are set and code paths are wired."
        )
    return schemas.IntegrationCatalogResponse(items=items, product_demo_mode=demo, demo_notice=notice)


def _decode_csv_rows(file_bytes: bytes) -> list[dict[str, str]]:
    text = file_bytes.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def _decode_spreadsheet_rows(filename: str, file_bytes: bytes) -> list[dict[str, str]]:
    if filename.lower().endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        parsed_rows: list[dict[str, str]] = []
        for row in rows[1:]:
            parsed_rows.append(
                {
                    headers[idx]: "" if value is None else str(value)
                    for idx, value in enumerate(row)
                    if idx < len(headers) and headers[idx]
                }
            )
        return parsed_rows
    return _decode_csv_rows(file_bytes)


def _import_deals_contacts(db: Session, rows: list[dict[str, str]]) -> int:
    imported = 0
    for row in rows:
        deal = (
            db.query(models.Deal)
            .filter(models.Deal.name == (row.get("deal_name") or "").strip())
            .first()
        )
        if not deal:
            deal = models.Deal(
                name=(row.get("deal_name") or "").strip(),
                description=row.get("description"),
                asset_type=row.get("asset_type"),
                city=row.get("city"),
                state=row.get("state"),
                source=row.get("source"),
                priority=(row.get("priority") or "medium").strip() or "medium",
                owner_name=row.get("owner_name"),
                next_action=row.get("next_action"),
            )
            db.add(deal)
            db.flush()
            db.add(
                models.DealStageEvent(
                    deal_id=deal.id,
                    from_stage=None,
                    to_stage=deal.stage,
                    reason="Imported from CSV",
                    author="importer",
                )
            )
        company = _get_or_create_company(
            db,
            company_name=row.get("company_name"),
            company_type=row.get("contact_type") or "investor",
            investor_type=row.get("investor_type"),
        )
        email = (row.get("contact_email") or "").strip()
        contact = db.query(models.Contact).filter(models.Contact.email == email).first() if email else None
        if not contact:
            contact = models.Contact(
                full_name=(row.get("contact_full_name") or "").strip(),
                email=email or None,
                contact_type=(row.get("contact_type") or "investor").strip() or "investor",
                company_id=company.id if company else None,
                investor_type=row.get("investor_type"),
                deal_id=deal.id,
            )
            db.add(contact)
        imported += 1
    db.commit()
    return imported


def _import_investor_pipeline(db: Session, rows: list[dict[str, str]]) -> int:
    imported = 0
    for row in rows:
        deal_name = (row.get("deal_name") or "").strip()
        email = (row.get("contact_email") or "").strip()
        deal = db.query(models.Deal).filter(models.Deal.name == deal_name).first()
        contact = db.query(models.Contact).filter(models.Contact.email == email).first() if email else None
        if not deal or not contact:
            continue
        entry = (
            db.query(models.InvestorPipelineEntry)
            .filter(
                models.InvestorPipelineEntry.deal_id == deal.id,
                models.InvestorPipelineEntry.contact_id == contact.id,
            )
            .first()
        )
        if not entry:
            entry = models.InvestorPipelineEntry(
                deal_id=deal.id,
                contact_id=contact.id,
            )
            db.add(entry)
        entry.status = (row.get("status") or "target").strip() or "target"
        entry.commitment_amount = int((row.get("commitment_amount") or "0").strip() or "0")
        entry.conviction = (row.get("conviction") or "medium").strip() or "medium"
        entry.last_signal = row.get("last_signal")
        entry.next_action = row.get("next_action")
        imported += 1
    db.commit()
    return imported


def _import_document_index(db: Session, rows: list[dict[str, str]]) -> int:
    imported = 0
    for row in rows:
        deal_name = (row.get("deal_name") or "").strip()
        deal = db.query(models.Deal).filter(models.Deal.name == deal_name).first()
        if not deal:
            continue
        document = models.Document(
            deal_id=deal.id,
            filename=(row.get("filename") or "imported_document.txt").strip(),
            document_type=(row.get("document_type") or "general").strip() or "general",
            status="processed",
            summary=row.get("summary"),
            risk_tags=row.get("risk_tags"),
            content=row.get("content") or row.get("summary") or "Imported document metadata row.",
        )
        db.add(document)
        imported += 1
    db.commit()
    return imported


@app.on_event("startup")
def startup() -> None:
    db = SessionLocal()
    try:
        create_default_admin(db)
    finally:
        db.close()


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.get("/health/ai")
def health_ai() -> dict:
    """Unauthenticated probe for local dev and smoke scripts; no secrets returned."""
    provider = current_ai_provider()
    ollama_ok: Optional[bool] = None
    if provider == "ollama":
        ollama_ok = ollama_server_reachable()
    return {
        "status": "ok",
        "ai_provider": provider,
        "ollama_reachable": ollama_ok,
        "local_fallback": os.getenv("REOS_AI_MODE", "").strip().lower() == "local_fallback",
        "predictive_ml": "not_deployed",
        "note": "Predictive ML training is a later phase; default stack is Ollama or local_fallback without external APIs.",
    }


@app.get("/auth/providers", response_model=schemas.AuthProvidersResponse)
def auth_providers(request: Request) -> schemas.AuthProvidersResponse:
    api_base = str(request.base_url).rstrip("/")
    return schemas.AuthProvidersResponse(**get_auth_providers(api_base))


@app.get("/auth/entra/start")
def auth_entra_start() -> RedirectResponse:
    if not microsoft_auth_configured():
        raise HTTPException(status_code=503, detail="Microsoft Entra is not configured for this environment")
    return RedirectResponse(build_entra_authorize_url())


@app.get("/auth/entra/callback")
async def auth_entra_callback(
    code: str | None = None, state: str | None = None, db: Session = Depends(get_db)
) -> RedirectResponse:
    if not code or not state:
        raise HTTPException(status_code=400, detail="Missing authorization code")
    expected_nonce = consume_entra_state(state)
    token_payload = await exchange_entra_code(code)
    id_token = token_payload.get("id_token")
    if not id_token:
        raise HTTPException(status_code=400, detail="Microsoft did not return an id_token")
    claims = await validate_entra_id_token(id_token, expected_nonce=expected_nonce)
    email = (claims.get("preferred_username") or claims.get("email") or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Microsoft account email is required")
    if not entra_user_allowed(email):
        raise HTTPException(status_code=403, detail="Microsoft account is not authorized for this workspace")
    username = email.split("@")[0]
    display_name = claims.get("name") or username
    role = map_entra_role(email)
    organization_name = (email.split("@", 1)[1] if "@" in email else "microsoft").lower()
    user = db.query(models.User).filter(models.User.email == email).first()
    if not user:
        user = models.User(
            username=username,
            email=email,
            password_hash=hash_password(secrets.token_urlsafe(24)),
            role=role,
            provider="microsoft",
            organization_name=organization_name,
            tenant_id=claims.get("tid"),
            display_name=display_name,
        )
        db.add(user)
        db.commit()
        db.refresh(user)
    else:
        user.provider = "microsoft"
        user.role = role
        user.display_name = display_name
        user.organization_name = organization_name
        user.tenant_id = claims.get("tid")
        db.commit()
        db.refresh(user)
    token = create_user_session(db, user, "microsoft")
    frontend_callback = os.getenv("REOS_ENTRA_FRONTEND_CALLBACK", "http://127.0.0.1:3001/auth/microsoft/callback")
    grant = issue_auth_grant(token)
    return RedirectResponse(f"{frontend_callback}?grant={grant}")


@app.post("/auth/login", response_model=schemas.LoginResponse)
def login(payload: schemas.LoginRequest, db: Session = Depends(get_db)) -> schemas.LoginResponse:
    if not local_recovery_enabled():
        raise HTTPException(status_code=403, detail="Local recovery login is disabled for this environment")
    user = db.query(models.User).filter(models.User.username == payload.username).first()
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_user_session(db, user, user.provider or "local")
    return schemas.LoginResponse(
        username=user.username,
        token=token,
        role=user.role,
        provider=user.provider or "local",
    )


@app.post("/auth/exchange", response_model=schemas.LoginResponse)
def auth_exchange(payload: schemas.AuthGrantExchangeRequest, db: Session = Depends(get_db)) -> schemas.LoginResponse:
    token = consume_auth_grant(payload.grant)
    identity = require_auth(db=db, authorization=f"Bearer {token}")
    username, role = identity
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return schemas.LoginResponse(username=username, token=token, role=role, provider=user.provider or "local")


@app.post("/auth/signup", response_model=schemas.SignupResponse)
def signup(payload: schemas.SignupRequest, db: Session = Depends(get_db)) -> schemas.SignupResponse:
    if not local_signup_enabled() or not local_recovery_enabled():
        raise HTTPException(
            status_code=403,
            detail="Direct signup is disabled. Provision access through Microsoft Entra or an administrator.",
        )
    username = (payload.username or "").strip()
    if len(username) < 3:
        raise HTTPException(status_code=400, detail="Username must be at least 3 characters")
    if not username.replace("_", "").isalnum():
        raise HTTPException(status_code=400, detail="Username may contain letters, numbers, and underscores only")
    raw_email = (payload.email or "").strip().lower()
    email = raw_email if raw_email else f"{username}@local.reos"
    if "@" not in email or len(email) < 5:
        raise HTTPException(status_code=400, detail="A valid email is required")

    if db.query(models.User).filter(models.User.username == username).first():
        raise HTTPException(status_code=409, detail="Username already registered")
    if db.query(models.User).filter(models.User.email == email).first():
        raise HTTPException(status_code=409, detail="Email already registered")

    password = payload.password or ""
    if len(password) < 10:
        raise HTTPException(status_code=400, detail="Password must be at least 10 characters")

    display_name = (payload.display_name or "").strip() or username.replace("_", " ").title()
    org = (payload.organization_name or "").strip() or None

    user = models.User(
        username=username,
        email=email,
        password_hash=hash_password(password),
        role="analyst",
        provider="local",
        display_name=display_name,
        organization_name=org,
    )
    db.add(user)
    db.flush()
    _record_audit_event(
        db,
        actor=username,
        action="local_signup",
        entity_type="user",
        entity_id=str(user.id),
        detail="Self-provisioned analyst account (local signup enabled).",
    )
    db.commit()
    return schemas.SignupResponse(
        username=username,
        role="analyst",
        message="Account created. Sign in with your new username and password.",
    )


@app.post("/auth/logout", response_model=schemas.LogoutResponse)
def auth_logout(
    db: Session = Depends(get_db),
    authorization: str = Header(default=""),
) -> schemas.LogoutResponse:
    token = authorization.replace("Bearer ", "", 1) if authorization.startswith("Bearer ") else ""
    if token:
        revoke_session(db, token)
    return schemas.LogoutResponse(message="Session closed.")


@app.get("/auth/me", response_model=schemas.MeResponse)
def auth_me(identity: tuple = Depends(require_auth)) -> schemas.MeResponse:
    username, role = identity
    return schemas.MeResponse(username=username, role=role, provider="session")


@app.post("/deals", response_model=schemas.DealOut)
def create_deal(
    payload: schemas.DealCreate,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> models.Deal:
    username, _ = identity
    deal = models.Deal(
        name=payload.name,
        description=payload.description,
        asset_type=payload.asset_type,
        city=payload.city,
        state=payload.state,
        source=payload.source,
        priority=payload.priority or "medium",
        owner_name=payload.owner_name or username,
        next_action=payload.next_action,
    )
    db.add(deal)
    db.flush()
    db.add(
        models.DealStageEvent(
            deal_id=deal.id,
            from_stage=None,
            to_stage=deal.stage,
            reason="Deal created",
            author=username,
        )
    )
    _record_audit_event(
        db,
        actor=username,
        action="create",
        entity_type="deal",
        entity_id=str(deal.id),
        detail=f"Created deal {deal.name}.",
    )
    db.commit()
    db.refresh(deal)
    return deal


@app.get("/deals", response_model=List[schemas.DealOut])
def list_deals(db: Session = Depends(get_db), _: tuple = Depends(require_auth)) -> List[models.Deal]:
    return db.query(models.Deal).order_by(models.Deal.id.desc()).all()


@app.patch("/deals/{deal_id}/stage", response_model=schemas.DealOut)
def update_stage(
    deal_id: int,
    payload: schemas.DealUpdateStage,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> models.Deal:
    username, _ = identity
    deal = db.query(models.Deal).filter(models.Deal.id == deal_id).first()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    previous_stage = deal.stage
    deal.stage = payload.stage
    db.add(
        models.DealStageEvent(
            deal_id=deal.id,
            from_stage=previous_stage,
            to_stage=payload.stage,
            reason="Stage updated from dashboard",
            author=username,
        )
    )
    _record_audit_event(
        db,
        actor=username,
        action="stage_update",
        entity_type="deal",
        entity_id=str(deal.id),
        detail=f"Moved from {previous_stage} to {payload.stage}.",
    )
    db.commit()
    db.refresh(deal)
    return deal


@app.post("/crm/contacts", response_model=schemas.ContactOut)
def create_contact(
    payload: schemas.ContactCreate,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> models.Contact:
    username, _ = identity
    company = _get_or_create_company(
        db,
        company_name=payload.company_name,
        company_type=payload.contact_type if payload.contact_type != "investor" else "investor",
        investor_type=payload.investor_type,
    )
    contact = models.Contact(
        full_name=payload.full_name,
        email=payload.email,
        contact_type=payload.contact_type,
        title=payload.title,
        phone=payload.phone,
        investor_type=payload.investor_type,
        company_id=company.id if company else None,
        deal_id=payload.deal_id,
        notes=payload.notes,
    )
    db.add(contact)
    _record_audit_event(
        db,
        actor=username,
        action="create",
        entity_type="contact",
        entity_id=str(contact.id or ""),
        detail=f"Created contact {payload.full_name}.",
    )
    db.commit()
    db.refresh(contact)
    return contact


@app.get("/crm/contacts", response_model=List[schemas.ContactOut])
def list_contacts(db: Session = Depends(get_db), _: tuple = Depends(require_auth)) -> List[models.Contact]:
    return (
        db.query(models.Contact)
        .options(joinedload(models.Contact.company))
        .order_by(models.Contact.id.desc())
        .limit(200)
        .all()
    )


@app.patch("/crm/contacts/{contact_id}", response_model=schemas.ContactOut)
def update_contact(
    contact_id: int,
    payload: schemas.ContactUpdate,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> models.Contact:
    username, _ = identity
    contact = db.query(models.Contact).filter(models.Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(contact, key, value)
    _record_audit_event(
        db,
        actor=username,
        action="update",
        entity_type="contact",
        entity_id=str(contact.id),
        detail="Updated contact fields from CRM.",
    )
    db.commit()
    db.refresh(contact)
    return contact


@app.get("/crm/companies", response_model=List[schemas.CompanyDetailOut])
def list_companies_detail(db: Session = Depends(get_db), _: tuple = Depends(require_auth)) -> List[schemas.CompanyDetailOut]:
    companies = db.query(models.Company).order_by(models.Company.id.desc()).limit(300).all()
    out: List[schemas.CompanyDetailOut] = []
    for c in companies:
        count = db.query(func.count(models.Contact.id)).filter(models.Contact.company_id == c.id).scalar() or 0
        out.append(
            schemas.CompanyDetailOut(
                id=c.id,
                name=c.name,
                company_type=c.company_type,
                investor_type=c.investor_type,
                notes=c.notes,
                created_at=c.created_at,
                contact_count=int(count),
            )
        )
    return out


@app.post("/crm/companies", response_model=schemas.CompanyOut)
def create_company(
    payload: schemas.CompanyCreate,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> models.Company:
    username, _ = identity
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Company name is required")
    if db.query(models.Company).filter(models.Company.name == name).first():
        raise HTTPException(status_code=409, detail="Company already exists")
    row = models.Company(
        name=name,
        company_type=payload.company_type or "investor",
        investor_type=payload.investor_type,
        notes=payload.notes,
    )
    db.add(row)
    db.flush()
    _record_audit_event(
        db,
        actor=username,
        action="create",
        entity_type="company",
        entity_id=str(row.id),
        detail=f"Created company {name}.",
    )
    db.commit()
    db.refresh(row)
    return row


@app.patch("/crm/companies/{company_id}", response_model=schemas.CompanyOut)
def update_company(
    company_id: int,
    payload: schemas.CompanyUpdate,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> models.Company:
    username, _ = identity
    company = db.query(models.Company).filter(models.Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(company, key, value)
    _record_audit_event(
        db,
        actor=username,
        action="update",
        entity_type="company",
        entity_id=str(company.id),
        detail="Updated company from CRM.",
    )
    db.commit()
    db.refresh(company)
    return company


@app.post("/crm/email-import/preview", response_model=schemas.EmailImportPreviewResponse)
def email_import_preview(
    payload: schemas.EmailImportPreviewRequest, _: tuple = Depends(require_auth)
) -> schemas.EmailImportPreviewResponse:
    detected_raw = parse_investor_email_text(payload.raw_text)
    detected = [
        schemas.EmailDetectedParty(
            email=p.email,
            full_name_guess=p.full_name_guess,
            company_guess=p.company_guess,
            decision_hint=p.decision_hint,
            rationale=p.rationale,
        )
        for p in detected_raw
    ]
    return schemas.EmailImportPreviewResponse(
        detected=detected,
        integration_note=(
            "Paste import only. For automatic mailbox ingestion, configure Microsoft Graph or Gmail in Integrations "
            "and run a server-side sync job; this endpoint does not open external mail APIs."
        ),
    )


@app.post("/crm/email-import/commit", response_model=schemas.EmailImportCommitResponse)
def email_import_commit(
    payload: schemas.EmailImportCommitRequest,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> schemas.EmailImportCommitResponse:
    username, _ = identity
    if payload.deal_id is not None:
        deal = db.query(models.Deal).filter(models.Deal.id == payload.deal_id).first()
        if not deal:
            raise HTTPException(status_code=404, detail="Deal not found")
    excerpt = (payload.body_excerpt or "").strip() or (payload.subject or "email import")[:1500]
    signals_created = 0
    contacts_created = 0
    contacts_matched = 0
    pipeline_updates = 0

    for item in payload.items:
        email_l = item.email.strip().lower()
        if not email_l:
            continue
        contact = (
            db.query(models.Contact)
            .filter(models.Contact.email.isnot(None))
            .filter(func.lower(models.Contact.email) == email_l)
            .first()
        )
        company_row: models.Company | None = None
        if contact:
            contacts_matched += 1
            company_row = contact.company
        elif payload.create_contacts:
            company = _get_or_create_company(
                db,
                company_name=item.company_name,
                company_type="investor",
                investor_type=item.investor_type,
            )
            contact = models.Contact(
                full_name=item.full_name.strip() or email_l.split("@")[0],
                email=email_l,
                contact_type="investor",
                investor_type=item.investor_type,
                company_id=company.id if company else None,
                deal_id=payload.deal_id,
                notes=f"Created from email import ({item.decision_hint}).",
            )
            db.add(contact)
            db.flush()
            company_row = company
            contacts_created += 1
        else:
            continue

        signal = models.InvestorEmailSignal(
            contact_id=contact.id,
            company_id=company_row.id if company_row else None,
            deal_id=payload.deal_id,
            sender_email=email_l,
            sender_name=item.full_name,
            subject_line=payload.subject,
            body_excerpt=excerpt[:8000],
            decision_inferred=item.decision_hint,
            source="paste",
            created_by=username,
        )
        db.add(signal)
        signals_created += 1

        if payload.apply_pipeline and payload.deal_id and contact:
            new_status = _pipeline_status_from_email_decision(item.decision_hint)
            entry = (
                db.query(models.InvestorPipelineEntry)
                .filter(
                    models.InvestorPipelineEntry.deal_id == payload.deal_id,
                    models.InvestorPipelineEntry.contact_id == contact.id,
                )
                .first()
            )
            if not entry:
                entry = models.InvestorPipelineEntry(
                    deal_id=payload.deal_id,
                    contact_id=contact.id,
                    status=new_status or "contacted",
                    last_signal=f"Email signal: {item.decision_hint}",
                )
                db.add(entry)
                pipeline_updates += 1
            else:
                if new_status:
                    entry.status = new_status
                note = f"Email signal: {item.decision_hint}"
                entry.last_signal = (entry.last_signal + " | " if entry.last_signal else "") + note
                pipeline_updates += 1

    db.commit()
    return schemas.EmailImportCommitResponse(
        signals_created=signals_created,
        contacts_created=contacts_created,
        contacts_matched=contacts_matched,
        pipeline_updates=pipeline_updates,
        message="Imported email-derived investor signals. Review pipeline and contacts for accuracy.",
    )


@app.get("/crm/email-signals", response_model=List[schemas.InvestorEmailSignalOut])
def list_email_signals(
    db: Session = Depends(get_db), _: tuple = Depends(require_auth)
) -> List[models.InvestorEmailSignal]:
    return (
        db.query(models.InvestorEmailSignal)
        .order_by(models.InvestorEmailSignal.id.desc())
        .limit(40)
        .all()
    )


@app.get("/documents/deal/{deal_id}", response_model=List[schemas.DocumentOut])
def list_deal_documents(
    deal_id: int,
    db: Session = Depends(get_db),
    _: tuple = Depends(require_auth),
) -> List[models.Document]:
    return (
        db.query(models.Document)
        .filter(models.Document.deal_id == deal_id)
        .order_by(models.Document.id.desc())
        .all()
    )


@app.post("/documents/{deal_id}/upload", response_model=schemas.DocumentOut)
async def upload_document(
    deal_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> models.Document:
    username, _ = identity
    deal = db.query(models.Deal).filter(models.Deal.id == deal_id).first()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    target = UPLOAD_DIR / file.filename
    with target.open("wb") as f:
        f.write(await file.read())

    try:
        content = extract_text(target).strip()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not content:
        raise HTTPException(status_code=400, detail="No extractable text from document")

    summary = " ".join(content.split()[:40]).strip()
    document = models.Document(
        deal_id=deal_id,
        filename=file.filename,
        document_type="uploaded_file",
        status="processed",
        summary=summary,
        risk_tags="review_required",
        content=content,
    )
    db.add(document)
    _record_audit_event(
        db,
        actor=username,
        action="upload",
        entity_type="document",
        entity_id=str(document.id or ""),
        detail=f"Uploaded {file.filename} to deal {deal_id}.",
    )
    db.commit()
    db.refresh(document)

    chunks = chunk_text(content)
    embeddings = []
    for c in chunks:
        embeddings.append((c, await embed_text(c)))
    store_chunks(db, deal_id, document.id, embeddings)
    return document


@app.post("/ai/query/{deal_id}", response_model=schemas.QueryResponse)
async def query_deal(
    deal_id: int,
    payload: schemas.QueryRequest,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> schemas.QueryResponse:
    username, _role = identity
    query_embedding = await embed_text(payload.question)
    chunks = retrieve_top_chunks(db, deal_id, query_embedding)
    if not chunks:
        return schemas.QueryResponse(answer="Insufficient evidence for this deal.", citations=[])
    answer = await generate_grounded_answer(payload.question, chunks)
    citations = [f"doc:{c.document_id}/chunk:{c.id}" for c in chunks]
    log = models.AIQueryLog(
        deal_id=deal_id,
        username=username,
        question=payload.question,
        answer=answer,
        citations=",".join(citations),
    )
    db.add(log)
    db.add(
        models.AIRun(
            workspace="deal_workspace",
            deal_id=deal_id,
            actor=username,
            provider=current_ai_provider(),
            model_name=current_ai_model(),
            prompt=payload.question,
            answer=answer,
            citations=",".join(citations),
            status="completed",
        )
    )
    _record_audit_event(
        db,
        actor=username,
        action="ai_query",
        entity_type="deal",
        entity_id=str(deal_id),
        detail=f"Ran grounded AI query with {len(citations)} citations.",
    )
    db.commit()
    return schemas.QueryResponse(answer=answer, citations=citations)


@app.get("/ai/history/{deal_id}", response_model=List[schemas.AIQueryLogOut])
def ai_history(
    deal_id: int,
    db: Session = Depends(get_db),
    _: tuple = Depends(require_auth),
) -> List[models.AIQueryLog]:
    return (
        db.query(models.AIQueryLog)
        .filter(models.AIQueryLog.deal_id == deal_id)
        .order_by(models.AIQueryLog.id.desc())
        .limit(20)
        .all()
    )


@app.post("/deals/{deal_id}/notes", response_model=schemas.DealNoteOut)
def add_deal_note(
    deal_id: int,
    payload: schemas.DealNoteCreate,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> models.DealNote:
    username, _ = identity
    deal = db.query(models.Deal).filter(models.Deal.id == deal_id).first()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    note = models.DealNote(deal_id=deal_id, author=username, content=payload.content.strip())
    db.add(note)
    _record_audit_event(
        db,
        actor=username,
        action="note",
        entity_type="deal",
        entity_id=str(deal_id),
        detail="Saved an internal note.",
    )
    db.commit()
    db.refresh(note)
    return note


@app.get("/deals/{deal_id}/notes", response_model=List[schemas.DealNoteOut])
def list_deal_notes(
    deal_id: int,
    db: Session = Depends(get_db),
    _: tuple = Depends(require_auth),
) -> List[models.DealNote]:
    return (
        db.query(models.DealNote)
        .filter(models.DealNote.deal_id == deal_id)
        .order_by(models.DealNote.id.desc())
        .limit(30)
        .all()
    )


@app.get("/analytics/summary", response_model=schemas.AnalyticsSummary)
def analytics_summary(db: Session = Depends(get_db), _: tuple = Depends(require_auth)) -> schemas.AnalyticsSummary:
    return _build_analytics_summary(db)


@app.post("/demo/seed", response_model=schemas.DemoSeedResponse)
def demo_seed(db: Session = Depends(get_db), _: tuple = Depends(require_roles("admin", "manager"))) -> schemas.DemoSeedResponse:
    response = _seed_demo_records(db)
    _ensure_workflow_records(db)
    added = _backfill_demo_dataset(db)
    return schemas.DemoSeedResponse(
        deals_created=response.deals_created,
        contacts_created=response.contacts_created,
        investor_pipeline_entries_created=response.investor_pipeline_entries_created,
        workflow_tasks_added=added["workflow_tasks_added"],
        workflow_exceptions_added=added["workflow_exceptions_added"],
        deal_notes_added=added["deal_notes_added"],
        documents_added=added["documents_added"],
        ai_query_logs_added=added["ai_query_logs_added"],
        ai_runs_added=added["ai_runs_added"],
        audit_events_added=added["audit_events_added"],
        chunks_added=added["chunks_added"],
    )


@app.get("/workspace/bootstrap", response_model=schemas.WorkspaceBootstrapResponse)
def workspace_bootstrap(
    db: Session = Depends(get_db), _: tuple = Depends(require_auth)
) -> schemas.WorkspaceBootstrapResponse:
    _ensure_workflow_records(db)
    deals = db.query(models.Deal).order_by(models.Deal.id.desc()).limit(25).all()
    contacts = db.query(models.Contact).order_by(models.Contact.id.desc()).limit(50).all()
    investor_pipeline = (
        db.query(models.InvestorPipelineEntry)
        .order_by(models.InvestorPipelineEntry.id.desc())
        .limit(50)
        .all()
    )
    return schemas.WorkspaceBootstrapResponse(
        analytics=_build_analytics_summary(db),
        deals=deals,
        contacts=contacts,
        investor_pipeline=investor_pipeline,
        operations=_operations_summary(db),
    )


@app.get("/deals/{deal_id}/workspace", response_model=schemas.DealWorkspaceResponse)
def deal_workspace(
    deal_id: int,
    db: Session = Depends(get_db),
    _: tuple = Depends(require_auth),
) -> schemas.DealWorkspaceResponse:
    deal = db.query(models.Deal).filter(models.Deal.id == deal_id).first()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    documents = (
        db.query(models.Document)
        .filter(models.Document.deal_id == deal_id)
        .order_by(models.Document.id.desc())
        .all()
    )
    diligence_items = (
        db.query(models.DiligenceItem)
        .filter(models.DiligenceItem.deal_id == deal_id)
        .order_by(models.DiligenceItem.id.desc())
        .all()
    )
    stage_events = (
        db.query(models.DealStageEvent)
        .filter(models.DealStageEvent.deal_id == deal_id)
        .order_by(models.DealStageEvent.id.desc())
        .all()
    )
    notes = (
        db.query(models.DealNote)
        .filter(models.DealNote.deal_id == deal_id)
        .order_by(models.DealNote.id.desc())
        .limit(30)
        .all()
    )
    ai_history = (
        db.query(models.AIQueryLog)
        .filter(models.AIQueryLog.deal_id == deal_id)
        .order_by(models.AIQueryLog.id.desc())
        .limit(20)
        .all()
    )
    investor_pipeline = (
        db.query(models.InvestorPipelineEntry)
        .filter(models.InvestorPipelineEntry.deal_id == deal_id)
        .order_by(models.InvestorPipelineEntry.id.desc())
        .all()
    )
    summary_lines = [
        f"Current stage: {deal.stage}",
        f"Priority: {deal.priority}",
        f"Next action: {deal.next_action or 'not set'}",
        f"Open diligence items: {sum(1 for item in diligence_items if item.status != 'done')}",
        f"Tracked investors: {len(investor_pipeline)}",
    ]
    decision_surface = _build_deal_decision_surface(db, deal, documents, diligence_items, investor_pipeline)
    return schemas.DealWorkspaceResponse(
        deal=deal,
        documents=documents,
        diligence_items=diligence_items,
        stage_events=stage_events,
        notes=notes,
        ai_history=ai_history,
        investor_pipeline=investor_pipeline,
        operations_summary=summary_lines,
        decision_surface=decision_surface,
    )


@app.post("/imports/csv", response_model=schemas.CsvImportResponse)
async def import_csv_file(
    import_type: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: tuple = Depends(require_auth),
) -> schemas.CsvImportResponse:
    rows = _decode_spreadsheet_rows(file.filename or "import.csv", await file.read())
    normalized_type = import_type.strip().lower()
    if normalized_type == "deals_contacts":
        imported = _import_deals_contacts(db, rows)
    elif normalized_type == "investor_pipeline":
        imported = _import_investor_pipeline(db, rows)
    elif normalized_type == "document_index":
        imported = _import_document_index(db, rows)
    else:
        raise HTTPException(status_code=400, detail="Unsupported import_type")
    return schemas.CsvImportResponse(import_type=normalized_type, rows_imported=imported)


@app.get("/integrations/catalog", response_model=schemas.IntegrationCatalogResponse)
def integrations_catalog(
    db: Session = Depends(get_db), _: tuple = Depends(require_auth)
) -> schemas.IntegrationCatalogResponse:
    return _build_integration_catalog(db)


@app.post("/integrations/catalog/toggle", response_model=schemas.IntegrationToggleResponse)
def toggle_integration_catalog_item(
    payload: schemas.IntegrationToggleRequest,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_roles("admin")),
) -> schemas.IntegrationToggleResponse:
    username, _ = identity
    valid_keys = {item["key"] for item in _integration_definitions(_integration_status_snapshot())}
    requested_key = payload.key.strip().lower()
    if requested_key not in valid_keys:
        raise HTTPException(status_code=404, detail="Integration not found")
    preference = db.query(models.IntegrationPreference).filter(models.IntegrationPreference.key == requested_key).first()
    if not preference:
        preference = models.IntegrationPreference(key=requested_key, enabled=payload.enabled)
        db.add(preference)
    else:
        preference.enabled = payload.enabled
    _record_audit_event(
        db,
        actor=username,
        action="integration_toggle",
        entity_type="integration",
        entity_id=requested_key,
        detail=f"Set enabled to {payload.enabled}.",
    )
    db.commit()
    return schemas.IntegrationToggleResponse(
        key=requested_key,
        enabled=payload.enabled,
        message="Integration toggle updated.",
    )


@app.get("/integrations/status", response_model=schemas.IntegrationStatus)
def integrations_status(_: tuple = Depends(require_auth)) -> schemas.IntegrationStatus:
    return _integration_status_snapshot()


@app.post("/integrations/mode", response_model=schemas.IntegrationModeUpdateResponse)
def update_integration_mode(
    payload: schemas.IntegrationModeUpdateRequest, _: tuple = Depends(require_roles("admin"))
) -> schemas.IntegrationModeUpdateResponse:
    requested = payload.mode.strip().lower()
    if requested not in {"local", "azure"}:
        raise HTTPException(status_code=400, detail="Mode must be either local or azure")
    os.environ["REOS_RUNTIME_MODE"] = requested
    if requested == "azure":
        os.environ["REOS_AI_PROVIDER"] = "azure_openai"
        message = "Azure runtime mode enabled. This is a config-level switch only."
    else:
        os.environ["REOS_AI_PROVIDER"] = "ollama"
        message = "Local runtime mode enabled. Azure integrations remain available for status checks."
    return schemas.IntegrationModeUpdateResponse(mode=requested, ai_provider=current_ai_provider(), message=message)


@app.get("/architecture/azure", response_model=schemas.AzureArchitectureResponse)
def azure_architecture(_: tuple = Depends(require_auth)) -> schemas.AzureArchitectureResponse:
    nodes = [
        schemas.AzureArchitectureNode(id="users", label="Users / Analysts", layer="edge", runtime="client"),
        schemas.AzureArchitectureNode(id="front_door", label="Azure Front Door", layer="edge", runtime="managed"),
        schemas.AzureArchitectureNode(id="app_gateway", label="Azure App Gateway", layer="edge", runtime="managed"),
        schemas.AzureArchitectureNode(id="api_mgmt", label="API Management", layer="api", runtime="managed"),
        schemas.AzureArchitectureNode(id="deal_service", label="Deal Service", layer="services", runtime="app_service"),
        schemas.AzureArchitectureNode(id="crm_service", label="CRM Service", layer="services", runtime="app_service"),
        schemas.AzureArchitectureNode(
            id="document_service", label="Document Service", layer="services", runtime="app_service"
        ),
        schemas.AzureArchitectureNode(id="blob_storage", label="Azure Blob Storage", layer="storage", runtime="managed"),
        schemas.AzureArchitectureNode(id="service_bus", label="Azure Service Bus", layer="messaging", runtime="managed"),
        schemas.AzureArchitectureNode(
            id="doc_processing", label="Document Processing", layer="processing", runtime="azure_functions"
        ),
        schemas.AzureArchitectureNode(
            id="embedding", label="Embedding Generation", layer="processing", runtime="azure_functions"
        ),
        schemas.AzureArchitectureNode(id="ai_search", label="Azure AI Search", layer="retrieval", runtime="managed"),
        schemas.AzureArchitectureNode(id="azure_openai", label="Azure OpenAI", layer="ai", runtime="managed"),
        schemas.AzureArchitectureNode(id="insights", label="AI Analysis Layer", layer="ai", runtime="app_service"),
    ]
    edges = [
        schemas.AzureArchitectureEdge(source="users", target="front_door", flow="HTTPS"),
        schemas.AzureArchitectureEdge(source="front_door", target="app_gateway", flow="WAF routing"),
        schemas.AzureArchitectureEdge(source="app_gateway", target="api_mgmt", flow="North-South API traffic"),
        schemas.AzureArchitectureEdge(source="api_mgmt", target="deal_service", flow="Deal API"),
        schemas.AzureArchitectureEdge(source="api_mgmt", target="crm_service", flow="CRM API"),
        schemas.AzureArchitectureEdge(source="api_mgmt", target="document_service", flow="Document API"),
        schemas.AzureArchitectureEdge(source="document_service", target="blob_storage", flow="Document write/read"),
        schemas.AzureArchitectureEdge(source="document_service", target="service_bus", flow="Ingestion event"),
        schemas.AzureArchitectureEdge(source="service_bus", target="doc_processing", flow="Queue trigger"),
        schemas.AzureArchitectureEdge(source="doc_processing", target="embedding", flow="Chunked text"),
        schemas.AzureArchitectureEdge(source="embedding", target="ai_search", flow="Vector upsert"),
        schemas.AzureArchitectureEdge(source="ai_search", target="azure_openai", flow="Retrieved context"),
        schemas.AzureArchitectureEdge(source="azure_openai", target="insights", flow="Grounded outputs"),
    ]
    return schemas.AzureArchitectureResponse(
        title="REOS Azure Enterprise Architecture",
        notes=[
            "This view models production topology while preserving local-first development.",
            "Mode switching is config-level and does not provision or tear down Azure resources.",
            "Service Bus and Functions isolate heavy asynchronous document processing.",
        ],
        nodes=nodes,
        edges=edges,
        deployment_stages=["Prototype", "Internal Testing", "Production"],
    )


@app.get("/automation/recommendations", response_model=schemas.AutomationRecommendationsResponse)
def automation_recommendations(
    db: Session = Depends(get_db), _: tuple = Depends(require_auth)
) -> schemas.AutomationRecommendationsResponse:
    total_deals = db.query(models.Deal).count()
    total_documents = db.query(models.Document).count()
    open_diligence = (
        db.query(models.Deal)
        .filter(models.Deal.stage.in_(["Screening", "Due Diligence", "Investment Committee"]))
        .count()
    )
    recommendations = [
        schemas.AutomationRecommendation(
            id="auto-risk-triage",
            title="Automate first-pass risk triage",
            impact="high",
            effort="medium",
            description="Run document-level risk tags immediately after upload and route high-risk deals to managers.",
            risk_if_ignored="Analysts spend time on low-risk files while critical issues wait in queue.",
        ),
        schemas.AutomationRecommendation(
            id="auto-stage-gates",
            title="Enforce stage gate checks automatically",
            impact="high",
            effort="low",
            description="Prevent stage transitions unless required artifacts and notes are present.",
            risk_if_ignored="Deals can progress without diligence evidence, creating audit and investment risk.",
        ),
        schemas.AutomationRecommendation(
            id="auto-digest",
            title="Send daily diligence digest",
            impact="medium",
            effort="low",
            description="Generate summary by owner: new docs, unresolved risks, and stage blockers.",
            risk_if_ignored="Leadership lacks consistent operational visibility and misses aging work.",
        ),
    ]
    challenges = [
        "Identity and RBAC mapping to existing Azure AD groups can cause permission drift if unmanaged.",
        "Model governance is required when mixing local Ollama and Azure OpenAI to avoid inconsistent outputs.",
        "Data residency and legal hold requirements may restrict where documents and embeddings can be stored.",
        "Automation without confidence thresholds can flood teams with false positives and alert fatigue.",
        "Legacy process dependencies in email/spreadsheets can block adoption unless workflow parity is planned.",
    ]
    if total_deals > 100 or total_documents > 300 or open_diligence > 40:
        recommendations.append(
            schemas.AutomationRecommendation(
                id="auto-capacity-balancing",
                title="Add workload balancing by analyst capacity",
                impact="high",
                effort="medium",
                description="Auto-assign incoming diligence tasks by queue depth, role, and SLA due dates.",
                risk_if_ignored="Backlog and turnaround time increase unevenly across teams.",
            )
        )
    return schemas.AutomationRecommendationsResponse(recommendations=recommendations, challenges=challenges)


@app.get("/portfolio/overview", response_model=schemas.PortfolioOverviewResponse)
def portfolio_overview(
    db: Session = Depends(get_db), identity: tuple = Depends(require_auth)
) -> schemas.PortfolioOverviewResponse:
    _username, _ = identity
    _ensure_workflow_records(db)
    analytics = _build_analytics_summary(db)
    active_deals = db.query(models.Deal).order_by(models.Deal.id.desc()).limit(8).all()
    committee_queue = [
        f"{deal.name} is in {deal.stage} with {deal.priority} priority and next action {deal.next_action or 'not set'}."
        for deal in active_deals
        if deal.stage in {"Investment Committee", "Approved", "Due Diligence"}
    ][:5]
    watchlist = [
        f"{deal.name} in {deal.city or 'Unknown'} is tracking source {deal.source or 'unassigned'}."
        for deal in active_deals[:5]
    ]
    investor_momentum = [
        f"{entry.status.title()} commitment of ${int(entry.commitment_amount or 0):,} with {entry.conviction or 'unknown'} conviction."
        for entry in db.query(models.InvestorPipelineEntry).order_by(models.InvestorPipelineEntry.id.desc()).limit(5).all()
    ]
    workflow_exceptions = [
        item.title for item in db.query(models.WorkflowException).order_by(models.WorkflowException.id.desc()).limit(5).all()
    ]
    ai_briefing = _read_only_briefing(
        committee_queue + investor_momentum + workflow_exceptions,
        "No portfolio briefing is available until operating records are loaded.",
    )
    return schemas.PortfolioOverviewResponse(
        analytics=analytics,
        stage_distribution=analytics.stage_distribution,
        committee_queue=committee_queue,
        watchlist=watchlist,
        investor_momentum=investor_momentum,
        workflow_exceptions=workflow_exceptions,
        ai_briefing=ai_briefing,
    )


def _apollo_live() -> bool:
    return bool(os.getenv("REOS_APOLLO_API_KEY", "").strip())


def _static_apollo_demo_feed() -> List[schemas.ApolloProspectOut]:
    return [
        schemas.ApolloProspectOut(
            person_name="Jordan Ellis",
            title="VP, Corporate Real Estate",
            organization="Northwind Logistics",
            fit_score=88,
            icp_match="Industrial / sale-leaseback",
            ai_note="Public expansion notes and two Sunbelt tours; stack aligns with logistics mandate.",
        ),
        schemas.ApolloProspectOut(
            person_name="Priya Shah",
            title="Director, Transactions",
            organization="Harbor Health Partners",
            fit_score=82,
            icp_match="Medical office",
            ai_note="Outpatient roll-up pattern matches Civic Medical Plaza thesis; warm intro path via lender.",
        ),
        schemas.ApolloProspectOut(
            person_name="Marcus Webb",
            title="Head of Acquisitions",
            organization="Crescent Family Office",
            fit_score=79,
            icp_match="Multifamily / core-plus",
            ai_note="Recent co-invest chatter on multifamily; timing overlaps Maple Grove screening.",
        ),
        schemas.ApolloProspectOut(
            person_name="Elena Varga",
            title="Managing Director",
            organization="Stonebank Lending",
            fit_score=74,
            icp_match="Debt / relationship",
            ai_note="Term sheet velocity strong on industrial; use for Harbor Logistics pacing.",
        ),
    ]


def _build_leads_ai_fit_ranked(db: Session) -> List[schemas.LeadsAiRankedDeal]:
    deals = (
        db.query(models.Deal)
        .filter(models.Deal.stage.in_(["Lead", "Screening"]))
        .order_by(models.Deal.id.desc())
        .limit(10)
        .all()
    )
    ranked: List[schemas.LeadsAiRankedDeal] = []
    priority_weight = {"high": 22, "medium": 10, "low": 0}
    asset_weight = {
        "industrial": 12,
        "multifamily": 10,
        "medical office": 9,
        "life sciences": 9,
        "office": 6,
        "retail": 5,
        "hospitality": 4,
        "self storage": 7,
        "mixed use": 5,
        "senior housing": 8,
    }
    for d in deals:
        base = 48
        score = base + priority_weight.get((d.priority or "medium").lower(), 0)
        at = (d.asset_type or "").lower()
        score += asset_weight.get(at, 6)
        if d.source and "broker" in d.source.lower():
            score += 5
        score = min(97, max(41, score))
        rationale = (
            f"{(d.asset_type or 'General').title()} in {d.city or 'TBD'}, {d.state or ''}: "
            f"priority {(d.priority or 'medium').lower()} with source {d.source or 'unspecified'}. "
            "Screen for sponsor track record and lease rollover before deep underwriting."
        )
        ranked.append(
            schemas.LeadsAiRankedDeal(
                deal_id=d.id,
                name=d.name,
                stage=d.stage,
                priority=d.priority,
                fit_score=score,
                rationale=rationale,
            )
        )
    ranked.sort(key=lambda r: (-r.fit_score, r.name))
    return ranked


def _deal_age_days(deal: models.Deal, *, now: datetime | None = None) -> int:
    reference = now or datetime.now(UTC)
    created = deal.created_at
    if created is None:
        return 0
    if created.tzinfo is None:
        created = created.replace(tzinfo=UTC)
    return max(0, int((reference - created).total_seconds() // 86400))


def _days_to_stage_entry(db: Session, target_stage: str) -> list[int]:
    """Calendar days from deal intake to first stage event hitting target_stage."""
    out: list[int] = []
    for deal in db.query(models.Deal).all():
        ev = (
            db.query(models.DealStageEvent)
            .filter(models.DealStageEvent.deal_id == deal.id, models.DealStageEvent.to_stage == target_stage)
            .order_by(models.DealStageEvent.created_at.asc())
            .first()
        )
        if not ev or not deal.created_at:
            continue
        start = deal.created_at.replace(tzinfo=UTC) if deal.created_at.tzinfo is None else deal.created_at
        end = ev.created_at.replace(tzinfo=UTC) if ev.created_at.tzinfo is None else ev.created_at
        out.append(max(0, int((end - start).total_seconds() // 86400)))
    return out


def _days_in_current_stage(db: Session, deal_id: int, current_stage: str) -> int | None:
    ev = (
        db.query(models.DealStageEvent)
        .filter(models.DealStageEvent.deal_id == deal_id, models.DealStageEvent.to_stage == current_stage)
        .order_by(models.DealStageEvent.created_at.desc())
        .first()
    )
    if not ev or not ev.created_at:
        return None
    start = ev.created_at.replace(tzinfo=UTC) if ev.created_at.tzinfo is None else ev.created_at
    return max(0, int((datetime.now(UTC) - start).total_seconds() // 86400))


def _build_deal_decision_surface(
    db: Session,
    deal: models.Deal,
    documents: list[models.Document],
    diligence_items: list[models.DiligenceItem],
    investor_pipeline: list[models.InvestorPipelineEntry],
) -> schemas.DealDecisionSurface:
    open_dil = [i for i in diligence_items if i.status != "done"]
    high_open = [i for i in open_dil if i.severity == "high"]
    age = _deal_age_days(deal)
    doc_n = len(documents)
    has_om = any(
        "om" in (d.filename or "").lower() or "offering" in (d.filename or "").lower() or "memorandum" in (d.filename or "").lower()
        for d in documents
    )
    has_model = any("model" in (d.filename or "").lower() or "xls" in (d.filename or "").lower() for d in documents)

    doc_checks = [
        f"{doc_n} documents in this deal workspace (target: OM + model + material contracts).",
        "Offering memo filename detected." if has_om else "No obvious offering memo filename; confirm OM is loaded.",
        "Financial model artifact present (by name)." if has_model else "No model file flagged by name; upload or link model.",
        "Use AI Q&A for clause and number cross-checks against your underwriting base case.",
    ]
    risks = [f"{i.title} [{i.severity}] — {i.status}" for i in open_dil[:6]]
    gaps: list[str] = []
    if doc_n == 0:
        gaps.append("No documents uploaded; IC cannot rely on corpus-backed answers.")
    if not deal.next_action:
        gaps.append("Next action is unset; no accountable forward motion.")
    if deal.stage in {"Lead", "Screening"} and age > 60:
        gaps.append(f"Deal age {age}d in early stage — screen to a decision or kill.")
    if high_open:
        gaps.append(f"{len(high_open)} high-severity diligence items still open.")

    committed = sum(1 for e in investor_pipeline if e.status == "committed")
    interested = sum(1 for e in investor_pipeline if e.status == "interested")
    passed = sum(1 for e in investor_pipeline if e.status == "passed")
    inv_summary = (
        f"Investor rows: {len(investor_pipeline)} ({committed} committed, {interested} interested, {passed} passed). "
        "Tie email-import signals to these rows for conversion tracking."
    )

    assumptions = [
        f"Stage gate: {deal.stage}",
        f"Priority label: {deal.priority}",
        "Rents, cap rate, and capex in any model are not re-computed here — human validation required.",
    ]
    if deal.city and deal.state:
        assumptions.append(f"Market: {deal.city}, {deal.state}")

    downside = (
        "Downside path if demand softens or tenant credit weakens: stretch exit and refi timeline; stress DSCR "
        "and sponsor liquidity before advancing IC."
    )

    actions: list[str] = []
    if gaps:
        actions.append("Clear the top blocking gap first (usually doc corpus or high-severity diligence).")
    if open_dil:
        actions.append(f"Close or downgrade {len(open_dil)} open diligence items with owners and dates.")
    if deal.stage == "Lead":
        actions.append("Move to Screening with written kill criteria or advance to diligence with IC sponsor.")
    if not has_om and doc_n > 0:
        actions.append("Tag or upload the OM so reviewers stop searching threadbare folders.")
    actions.append("Run a document-backed AI question focused on the single risk that would flip the vote.")
    if not actions:
        actions.append("Confirm IC materials and investor coverage; press toward vote or documented pass.")

    verdict = "insufficient_data"
    confidence = 42
    rationale = "Heuristic draft from stage, diligence, documents, and age — not a substitute for committee judgment."

    if deal.stage == "Passed":
        verdict = "reject"
        confidence = 88
        rationale = "Deal marked passed in system."
    elif deal.stage in {"Approved", "Closing"}:
        verdict = "proceed_execute"
        confidence = 82
        rationale = "Deal in execution-track stages; focus is closing conditions, not re-underwriting from scratch."
    elif doc_n == 0 and deal.stage not in {"Lead"}:
        verdict = "insufficient_data"
        confidence = 38
        rationale = "Diligence lane without corpus — block or force document ingest before IC narrative."
    elif len(high_open) >= 2:
        verdict = "hold"
        confidence = 58
        rationale = "Multiple high-severity open items — decision quality risk if forced now."
    elif deal.stage in {"Investment Committee"}:
        verdict = "ready_for_vote"
        confidence = 67 if not high_open else 52
        rationale = "In IC lane; verdict assumes packet aligns with open exceptions list."
    elif deal.stage in {"Due Diligence"}:
        verdict = "hold_advance"
        confidence = 60
        rationale = "Active diligence — advance when risks are bounded or kill with written rationale."
    elif deal.stage in {"Screening"} and age > 45:
        verdict = "kill_or_advance"
        confidence = 49
        rationale = "Long screening without progression — force a screen decision."
    elif deal.stage == "Lead":
        verdict = "screen"
        confidence = 45
        rationale = "Too early to treat as investment-ready; qualify or exit quickly."

    auto_note = (
        "Automation posture: this surface is rule-driven today. Next step is ML on outcomes (pass/loss/win) "
        "and structured extraction (leases, covenants) feeding the same verdict block with human sign-off."
    )

    return schemas.DealDecisionSurface(
        current_verdict=verdict,
        confidence=confidence,
        confidence_rationale=rationale,
        key_assumptions=assumptions,
        downside_scenario=downside,
        top_risks=risks or ["No tracked diligence risks — confirm checklist coverage."],
        blocking_gaps=gaps or ["No blocking gaps flagged by rules — still verify IC readiness."],
        next_best_actions=actions[:8],
        document_dd_checks=doc_checks,
        investor_posture_summary=inv_summary,
        automation_note=auto_note,
    )


def _investor_conversion_hints(db: Session) -> list[schemas.InvestorActionHint]:
    hints: list[schemas.InvestorActionHint] = []
    pipeline = (
        db.query(models.InvestorPipelineEntry)
        .order_by(models.InvestorPipelineEntry.id.desc())
        .limit(48)
        .all()
    )
    for entry in pipeline:
        st = (entry.status or "").lower()
        urgency = "medium"
        hint = ""
        if st in {"target", "contacted"} and not (entry.last_signal or "").strip():
            hint = "No logged reply — send a concrete ask (data room, IC date, ticket size) and paste the thread into CRM import."
            urgency = "high"
        elif st == "interested" and not (entry.next_action or "").strip():
            hint = "Interested but no next action — schedule IC preview or soft circle and record outcome."
            urgency = "high"
        elif st == "interested":
            hint = "Maintain momentum: confirm diligence depth matches their mandate and update last_signal after each touch."
            urgency = "medium"
        elif st == "committed":
            hint = "Committed — verify subscription docs and wire timing; keep last_signal current."
            urgency = "low"
        elif st == "passed":
            continue
        else:
            hint = "Review fit and either warm the relationship with a clear story or mark passed with reason."
            urgency = "low"
        if hint:
            hints.append(
                schemas.InvestorActionHint(
                    pipeline_entry_id=entry.id,
                    deal_id=entry.deal_id,
                    hint=hint,
                    urgency=urgency,
                )
            )
    return hints[:24]


def _median_int(values: list[int]) -> float:
    if not values:
        return 0.0
    s = sorted(values)
    mid = len(s) // 2
    if len(s) % 2:
        return float(s[mid])
    return (s[mid - 1] + s[mid]) / 2.0


def _build_operating_capability_rows(db: Session) -> list[schemas.OperatingCapabilityRow]:
    leadish = (
        db.query(models.Deal)
        .filter(models.Deal.stage.in_(["Lead", "Screening"]))
        .count()
    )
    doc_total = db.query(models.Document).count()
    contact_total = db.query(models.Contact).count()
    company_total = db.query(models.Company).count()
    inv_pipe = db.query(models.InvestorPipelineEntry).count()
    sig_total = db.query(models.InvestorEmailSignal).count()
    return [
        schemas.OperatingCapabilityRow(
            id="find_deals",
            problem="Finding deals",
            status="ready",
            route_path="/app/leads",
            route_label="Intake & mandate fit",
            detail=f"Origination queue and ranked fit preview. {leadish} deals in Lead or Screening.",
        ),
        schemas.OperatingCapabilityRow(
            id="contacts",
            problem="Organising contacts",
            status="ready",
            route_path="/app/crm",
            route_label="CRM & companies",
            detail=f"{contact_total} contacts across {company_total} companies — notes, investor type, email-import signals.",
        ),
        schemas.OperatingCapabilityRow(
            id="stages",
            problem="Tracking deal stages",
            status="ready",
            route_path="/app/deals",
            route_label="Deals hub & workspace",
            detail="Per-deal timeline and stage control from the workspace (audited transitions).",
        ),
        schemas.OperatingCapabilityRow(
            id="documents",
            problem="Reviewing large document sets",
            status="ready",
            route_path="/app/documents",
            route_label="Document library",
            detail=f"{doc_total} files in library; per-deal upload plus citation-backed Q&A in deal workspace.",
        ),
        schemas.OperatingCapabilityRow(
            id="investors",
            problem="Chasing investors",
            status="ready",
            route_path="/app/investors",
            route_label="Pipeline & onboarding",
            detail=f"{inv_pipe} investor pipeline rows tied to contacts; CRM email import logged {sig_total} times.",
        ),
        schemas.OperatingCapabilityRow(
            id="reports",
            problem="Building reports manually",
            status="partial",
            route_path="/app/reports",
            route_label="Committee & reporting queue",
            detail="Standing brief and queue here; one-click IC packets and branded exports still to ship.",
        ),
    ]


def _build_decision_velocity(db: Session) -> schemas.DecisionVelocitySummary:
    now = datetime.now(UTC)
    active_stages = frozenset({"Lead", "Screening", "Due Diligence", "Investment Committee", "Approved", "Closing"})
    deals = db.query(models.Deal).filter(models.Deal.stage.in_(active_stages)).all()
    ages: list[int] = []
    ic_ages: list[int] = []
    for d in deals:
        created = d.created_at
        if created is None:
            continue
        if created.tzinfo is None:
            created = created.replace(tzinfo=UTC)
        days = max(0, int((now - created).total_seconds() // 86400))
        ages.append(days)
        if d.stage in {"Due Diligence", "Investment Committee", "Approved", "Closing"}:
            ic_ages.append(days)
    med_all = int(round(_median_int(ages)))
    med_ic = int(round(_median_int(ic_ages)))

    doc_total = db.query(models.Document).count()
    deal_ids_with_docs = (
        db.query(func.count(func.distinct(models.Document.deal_id))).scalar() or 0
    )
    open_deal_count = max(len(deals), 1)
    doc_intensity = round(doc_total / open_deal_count, 1)

    pipeline = db.query(models.InvestorPipelineEntry).count()
    with_signal = (
        db.query(models.InvestorPipelineEntry)
        .filter(models.InvestorPipelineEntry.last_signal.isnot(None))
        .filter(models.InvestorPipelineEntry.last_signal != "")
        .count()
    )

    to_dd = _days_to_stage_entry(db, "Due Diligence")
    median_dd = int(round(_median_int(to_dd))) if to_dd else 0
    ic_lane: list[int] = []
    for d_row in db.query(models.Deal).filter(models.Deal.stage == "Investment Committee").all():
        span = _days_in_current_stage(db, d_row.id, "Investment Committee")
        if span is not None:
            ic_lane.append(span)
    median_ic_dwell = int(round(_median_int(ic_lane))) if ic_lane else 0

    return schemas.DecisionVelocitySummary(
        headline="Median days in pipeline (open deals)",
        primary_value=f"{med_all} days",
        primary_subtext=(
            f"Deals in diligence+ lanes average {med_ic} days since intake. "
            f"Median calendar days intake to first Due Diligence event: {median_dd if median_dd else 'n/a'}."
        ),
        metrics=[
            schemas.DecisionVelocityMetric(label="Distinct deals with docs", value=str(deal_ids_with_docs)),
            schemas.DecisionVelocityMetric(label="Documents per open deal", value=str(doc_intensity)),
            schemas.DecisionVelocityMetric(
                label="Investor rows with last signal text", value=f"{with_signal} / {pipeline}" if pipeline else "0 / 0"
            ),
        ],
        median_days_to_diligence=f"{median_dd} days" if median_dd else "n/a (no DD transitions logged)",
        median_days_in_investment_committee=(
            f"{median_ic_dwell} days in IC lane" if median_ic_dwell else "n/a (no deals currently in IC)"
        ),
        methodology_note=(
            "Intake time uses deal created_at. Days-to-diligence uses first stage event to Due Diligence. "
            "IC dwell uses days since the stage event that moved the deal into Investment Committee. "
            "All metrics are operational proxies until broker timestamps and board dates are first-class fields."
        ),
    )


def _dashboard_seed_data(db: Session) -> schemas.DashboardDataResponse:
    """KPIs and widgets for the overview; strategy narrative lives under /app/strategy."""
    deal_total = db.query(models.Deal).count()
    origination = db.query(models.Deal).filter(models.Deal.stage.in_(["Lead", "Screening"])).count()
    diligence = db.query(models.Deal).filter(models.Deal.stage.in_(["Due Diligence", "Investment Committee"])).count()
    investor_rows = db.query(models.InvestorPipelineEntry).count()
    committed = db.query(models.InvestorPipelineEntry).filter(models.InvestorPipelineEntry.status == "committed").count()

    kpis = [
        schemas.DashboardKpi(
            label="Active deals",
            value=str(deal_total) if deal_total else "0",
            change="In workspace",
            change_positive=True,
        ),
        schemas.DashboardKpi(
            label="Diligence / IC",
            value=str(diligence) if diligence else "0",
            change="Committee attention",
            change_positive=diligence <= 6,
        ),
        schemas.DashboardKpi(
            label="Origination queue",
            value=str(origination) if origination else "0",
            change="Lead + screening",
            change_positive=True,
        ),
        schemas.DashboardKpi(
            label="Investor pipeline",
            value=str(investor_rows) if investor_rows else "0",
            change=f"{committed} committed" if committed else "Open entries",
            change_positive=True,
        ),
    ]
    revenue_forecast = {
        "value": "—",
        "change": "Set in reporting",
        "label": "Deploy / capital (internal)",
    }
    market_trends = [
        schemas.DashboardMarketPoint(month="Feb", median_sale_price=892000, inventory=420),
        schemas.DashboardMarketPoint(month="Mar", median_sale_price=905000, inventory=398),
        schemas.DashboardMarketPoint(month="May", median_sale_price=918000, inventory=385),
        schemas.DashboardMarketPoint(month="Jul", median_sale_price=932000, inventory=372),
        schemas.DashboardMarketPoint(month="Sep", median_sale_price=945000, inventory=358),
        schemas.DashboardMarketPoint(month="Nov", median_sale_price=962000, inventory=340),
    ]
    tasks = [
        schemas.DashboardTask(assignee="IC Secretary", description="Packet for Harbor Logistics vote", time_ago="Today"),
        schemas.DashboardTask(assignee="Capital Markets", description="LP questions on West Loop rent roll", time_ago="3h"),
        schemas.DashboardTask(assignee="Legal", description="Title exceptions on Maple Grove", time_ago="Yesterday"),
        schemas.DashboardTask(assignee="Asset Mgmt", description="Capex bids for Riverfront PIP", time_ago="2d"),
    ]
    pipe_counts = [
        ("Lead", db.query(models.Deal).filter(models.Deal.stage == "Lead").count() or 3),
        ("Screening", db.query(models.Deal).filter(models.Deal.stage == "Screening").count() or 4),
        ("Due Diligence", db.query(models.Deal).filter(models.Deal.stage == "Due Diligence").count() or 3),
        ("IC / Approved", db.query(models.Deal).filter(models.Deal.stage.in_(["Investment Committee", "Approved"])).count() or 2),
    ]
    colors = ["#1d4ed8", "#3b82f6", "#0d9488", "#16a34a"]
    sales_pipeline = [
        schemas.DashboardPipelineStage(
            name=name,
            count=max(1, cnt),
            value=float(max(1, cnt) * 875000),
            color=colors[i % len(colors)],
        )
        for i, (name, cnt) in enumerate(pipe_counts)
    ]
    top_properties = [
        schemas.DashboardTopProperty(address="West Loop Office, Chicago, IL", agent="Acquisitions East", price=118000000),
        schemas.DashboardTopProperty(address="Harbor Logistics Park, Savannah, GA", agent="Industrial", price=89500000),
        schemas.DashboardTopProperty(address="Maple Grove Multifamily, Dallas, TX", agent="Sunbelt", price=62400000),
    ]
    top_agents = [
        schemas.DashboardTopAgent(name="Acquisitions East", sales_count=6, total_sales=410000000),
        schemas.DashboardTopAgent(name="Industrial", sales_count=5, total_sales=382000000),
        schemas.DashboardTopAgent(name="Sunbelt", sales_count=7, total_sales=355000000),
    ]
    activity_feed = [
        schemas.DashboardActivity(title="Review open diligence items", due="Today"),
        schemas.DashboardActivity(title="Investor data room follow-ups", due="Today"),
        schemas.DashboardActivity(title="Committee materials refresh", due="This week"),
    ]

    perf = min(96, 72 + min(deal_total, 12) * 2)

    return schemas.DashboardDataResponse(
        kpis=kpis,
        revenue_forecast=revenue_forecast,
        market_trends=market_trends,
        tasks=tasks,
        sales_pipeline=sales_pipeline,
        top_properties=top_properties,
        top_agents=top_agents,
        activity_feed=activity_feed,
        performance_score=perf,
        performance_max=100,
        performance_label="Operating rhythm",
        performance_subtitle="Based on deal volume in workspace",
        operating_capabilities=_build_operating_capability_rows(db),
        decision_velocity=_build_decision_velocity(db),
    )


@app.get("/dashboard/data", response_model=schemas.DashboardDataResponse)
def dashboard_data(db: Session = Depends(get_db), _: tuple = Depends(require_auth)) -> schemas.DashboardDataResponse:
    return _dashboard_seed_data(db)


PLAYBOOK_GROUP_KEYS = frozenset(
    {"1-5", "6-10", "11-18", "19-26", "27-31", "32-36", "37-41", "42-46", "47-50"}
)


def _user_id_from_username(db: Session, username: str) -> int:
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user.id


def _build_executive_briefing(db: Session) -> List[str]:
    _ensure_workflow_records(db)
    lines: List[str] = []
    deals = db.query(models.Deal).order_by(models.Deal.id.desc()).limit(18).all()
    highs = [d for d in deals if d.priority == "high"]
    if highs:
        lines.append(f"{len(highs)} high-priority deals need attention this week.")
        for d in highs[:4]:
            lines.append(f"{d.name} | {d.stage} | Next: {d.next_action or 'set next action'}")
    pipe_count = 0
    for d in deals:
        if d.stage in {"Investment Committee", "Approved", "Due Diligence", "Closing", "Screening"}:
            lines.append(f"Active pipeline: {d.name} ({d.stage}, {d.priority})")
            pipe_count += 1
            if pipe_count >= 4:
                break
    for ex in (
        db.query(models.WorkflowException)
        .filter(models.WorkflowException.status == "open")
        .order_by(models.WorkflowException.id.desc())
        .limit(5)
        .all()
    ):
        lines.append(f"Open exception: {ex.title}")
    for item in (
        db.query(models.DiligenceItem)
        .filter(models.DiligenceItem.status != "done")
        .order_by(models.DiligenceItem.id.desc())
        .limit(4)
        .all()
    ):
        lines.append(f"Diligence: {item.title} on deal #{item.deal_id} ({item.status}, {item.severity})")
    for entry in (
        db.query(models.InvestorPipelineEntry).order_by(models.InvestorPipelineEntry.id.desc()).limit(4).all()
    ):
        contact = entry.contact
        cname = contact.full_name if contact else f"Contact #{entry.contact_id}"
        lines.append(f"Investor momentum: {cname} — {entry.status}, ${int(entry.commitment_amount or 0):,}")
    if not lines:
        lines.append("No briefing signals yet. Run demo seed or add deals, diligence, and investor records.")
    return lines[:16]


@app.get("/executive/briefing", response_model=schemas.ExecutiveBriefingResponse)
def executive_briefing(db: Session = Depends(get_db), _: tuple = Depends(require_auth)) -> schemas.ExecutiveBriefingResponse:
    return schemas.ExecutiveBriefingResponse(lines=_build_executive_briefing(db))


def _playbook_checklist_response(db: Session, user_id: int) -> schemas.PlaybookChecklistResponse:
    rows = db.query(models.PlaybookChecklistItem).filter(models.PlaybookChecklistItem.user_id == user_id).all()
    by_key = {r.group_key: r.complete for r in rows}
    items = [schemas.PlaybookChecklistRow(group_key=k, complete=by_key.get(k, False)) for k in sorted(PLAYBOOK_GROUP_KEYS)]
    return schemas.PlaybookChecklistResponse(items=items)


@app.get("/playbook/checklist", response_model=schemas.PlaybookChecklistResponse)
def playbook_checklist(
    db: Session = Depends(get_db), identity: tuple = Depends(require_auth)
) -> schemas.PlaybookChecklistResponse:
    username, _ = identity
    uid = _user_id_from_username(db, username)
    return _playbook_checklist_response(db, uid)


@app.post("/playbook/checklist/toggle", response_model=schemas.PlaybookChecklistResponse)
def playbook_checklist_toggle(
    payload: schemas.PlaybookChecklistToggleRequest,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> schemas.PlaybookChecklistResponse:
    if payload.group_key not in PLAYBOOK_GROUP_KEYS:
        raise HTTPException(status_code=400, detail="Unknown automation group key")
    username, _ = identity
    uid = _user_id_from_username(db, username)
    row = (
        db.query(models.PlaybookChecklistItem)
        .filter(
            models.PlaybookChecklistItem.user_id == uid,
            models.PlaybookChecklistItem.group_key == payload.group_key,
        )
        .first()
    )
    if row:
        row.complete = not row.complete
        row.updated_at = datetime.now(UTC)
    else:
        db.add(models.PlaybookChecklistItem(user_id=uid, group_key=payload.group_key, complete=True))
    db.commit()
    return _playbook_checklist_response(db, uid)


@app.get("/crm/graph", response_model=schemas.CRMGraphResponse)
def crm_graph(db: Session = Depends(get_db), _: tuple = Depends(require_auth)) -> schemas.CRMGraphResponse:
    nodes: List[schemas.CRMGraphNode] = []
    edges: List[schemas.CRMGraphEdge] = []
    seen_ids: set[str] = set()

    contacts = db.query(models.Contact).order_by(models.Contact.id.desc()).limit(24).all()
    company_ids_from_contacts = {c.company_id for c in contacts if c.company_id}
    companies = db.query(models.Company).order_by(models.Company.id.desc()).limit(14).all()
    company_ids_seen = {c.id for c in companies}
    missing_company_ids = company_ids_from_contacts - company_ids_seen
    if missing_company_ids:
        extra = db.query(models.Company).filter(models.Company.id.in_(list(missing_company_ids))).all()
        companies = list(companies) + list(extra)

    for c in companies:
        nid = f"company-{c.id}"
        seen_ids.add(nid)
        nodes.append(schemas.CRMGraphNode(id=nid, label=c.name[:48], node_type="company"))

    deal_ids_needed = {c.deal_id for c in contacts if c.deal_id}
    deals_by_id: Dict[int, models.Deal] = {}
    if deal_ids_needed:
        for d in db.query(models.Deal).filter(models.Deal.id.in_(list(deal_ids_needed))).all():
            deals_by_id[d.id] = d

    for c in contacts:
        pid = f"contact-{c.id}"
        if pid not in seen_ids:
            seen_ids.add(pid)
            nodes.append(schemas.CRMGraphNode(id=pid, label=c.full_name[:40], node_type="contact"))
        if c.company_id:
            cid = f"company-{c.company_id}"
            if cid in seen_ids:
                edges.append(schemas.CRMGraphEdge(source=pid, target=cid, relation="at_company"))
        if c.deal_id and c.deal_id in deals_by_id:
            did = f"deal-{c.deal_id}"
            if did not in seen_ids:
                seen_ids.add(did)
                d = deals_by_id[c.deal_id]
                nodes.append(schemas.CRMGraphNode(id=did, label=d.name[:40], node_type="deal"))
            edges.append(schemas.CRMGraphEdge(source=pid, target=did, relation="on_deal"))

    return schemas.CRMGraphResponse(nodes=nodes, edges=edges)


@app.get("/deals/overview", response_model=schemas.DealsOverviewResponse)
def deals_overview(db: Session = Depends(get_db), _: tuple = Depends(require_auth)) -> schemas.DealsOverviewResponse:
    analytics = _build_analytics_summary(db)
    deals = db.query(models.Deal).order_by(models.Deal.id.desc()).all()
    intake_summary = [
        f"{deal.name} entered via {deal.source or 'unspecified'} and is owned by {deal.owner_name or 'unassigned'}."
        for deal in deals[:6]
    ]
    active_diligence = [deal for deal in deals if deal.stage in {"Screening", "Due Diligence", "Investment Committee"}][:12]
    closing_pipeline = [deal for deal in deals if deal.stage in {"Approved", "Closing"}][:8]
    return schemas.DealsOverviewResponse(
        analytics=analytics,
        intake_summary=intake_summary,
        active_diligence=active_diligence,
        closing_pipeline=closing_pipeline,
    )


@app.get("/leads/overview", response_model=schemas.LeadsOverviewResponse)
def leads_overview(
    db: Session = Depends(get_db), identity: tuple = Depends(require_auth)
) -> schemas.LeadsOverviewResponse:
    _username, _ = identity
    lead_deals = (
        db.query(models.Deal)
        .filter(models.Deal.stage.in_(["Lead", "Screening"]))
        .order_by(models.Deal.id.desc())
        .limit(12)
        .all()
    )
    origination_signals = [
        f"{deal.name} came from {deal.source or 'direct channel'} and is centered on {deal.asset_type or 'unspecified'}."
        for deal in lead_deals[:6]
    ]
    contacts = (
        db.query(models.Contact)
        .options(joinedload(models.Contact.company))
        .order_by(models.Contact.id.desc())
        .limit(6)
        .all()
    )
    outreach_queue = [
        f"{contact.full_name} at {(contact.company.name if contact.company else 'unassigned company')} needs outreach follow-up."
        for contact in contacts
    ]
    ai_briefing = _read_only_briefing(
        origination_signals + outreach_queue,
        "Lead coverage will appear once origination records are present.",
    )
    apollo = schemas.LeadsApolloConnectorStub(
        status="configured" if _apollo_live() else "demo_placeholder",
        last_sync_display="Never (demo)" if not _apollo_live() else "Scheduled / last run redacted",
        api_env_key="REOS_APOLLO_API_KEY",
        narrative=(
            "Live Apollo outbound calls are disabled until an API key is set and a sync worker is deployed. "
            "The UI shows representative prospects and AI fit ranking over workspace deals."
        ),
    )
    prospects = _static_apollo_demo_feed()
    return schemas.LeadsOverviewResponse(
        lead_deals=lead_deals,
        origination_signals=origination_signals,
        outreach_queue=outreach_queue,
        ai_briefing=ai_briefing,
        apollo=apollo,
        apollo_prospects=prospects,
    )


@app.get("/leads/ai-fit-preview", response_model=schemas.LeadsAiFitPreviewResponse)
def leads_ai_fit_preview(
    db: Session = Depends(get_db), _: tuple = Depends(require_auth)
) -> schemas.LeadsAiFitPreviewResponse:
    ranked = _build_leads_ai_fit_ranked(db)
    summary = (
        "Heuristic demo ranking over Lead and Screening deals (mandate fit proxy). "
        "Wire REOS to your LLM provider to add narrative variance and doc-aware features."
    )
    if ranked:
        top = ranked[0]
        summary = f"Top pick: {top.name} (score {top.fit_score}). {summary}"
    return schemas.LeadsAiFitPreviewResponse(
        generated_at=datetime.now(UTC).isoformat(),
        model_route="heuristic_demo_v1",
        ranked=ranked,
        summary=summary,
    )


@app.get("/crm/overview", response_model=schemas.CRMOverviewResponse)
def crm_overview(
    db: Session = Depends(get_db), identity: tuple = Depends(require_auth)
) -> schemas.CRMOverviewResponse:
    _username, _ = identity
    contacts = (
        db.query(models.Contact)
        .options(joinedload(models.Contact.company))
        .order_by(models.Contact.id.desc())
        .limit(40)
        .all()
    )
    companies = db.query(models.Company).order_by(models.Company.id.desc()).limit(40).all()
    contact_rows = [_contact_to_crm_row(c) for c in contacts]
    relationship_signals = [
        f"{contact.full_name} is a {contact.contact_type} linked to "
        f"{(contact.company_name or 'direct relationship')}."
        for contact in contact_rows[:8]
    ]
    ai_briefing = _read_only_briefing(
        relationship_signals,
        "Relationship coverage becomes clearer once contacts and companies are linked to active work.",
    )
    return schemas.CRMOverviewResponse(
        contacts=contact_rows,
        companies=companies,
        relationship_signals=relationship_signals,
        ai_briefing=ai_briefing,
        email_integration_hint=(
            "Pasted email import is live; Microsoft Graph and Gmail sync remain configuration-driven in Integrations."
        ),
    )


@app.get("/investors/overview", response_model=schemas.InvestorsOverviewResponse)
def investors_overview(
    db: Session = Depends(get_db), identity: tuple = Depends(require_auth)
) -> schemas.InvestorsOverviewResponse:
    _username, _ = identity
    pipeline = db.query(models.InvestorPipelineEntry).order_by(models.InvestorPipelineEntry.id.desc()).limit(24).all()
    focus_notes = [
        f"{entry.status.title()} at ${int(entry.commitment_amount or 0):,} with {entry.conviction or 'unknown'} conviction."
        for entry in pipeline[:8]
    ]
    onboarding_status = [
        f"{entry.next_action or 'Set onboarding step'} for investor entry #{entry.id}."
        for entry in pipeline[:8]
    ]
    ai_briefing = _read_only_briefing(
        focus_notes + onboarding_status,
        "Investor coverage will appear once commitment records begin to move.",
    )
    return schemas.InvestorsOverviewResponse(
        investor_pipeline=pipeline,
        focus_notes=focus_notes,
        onboarding_status=onboarding_status,
        ai_briefing=ai_briefing,
        conversion_action_hints=_investor_conversion_hints(db),
    )


@app.get("/documents/library", response_model=schemas.DocumentsLibraryResponse)
def documents_library(
    db: Session = Depends(get_db), identity: tuple = Depends(require_auth)
) -> schemas.DocumentsLibraryResponse:
    _username, _ = identity
    documents = db.query(models.Document).order_by(models.Document.id.desc()).limit(30).all()
    document_signals = [
        f"{document.filename} is {document.status} with risk tags {document.risk_tags or 'none'}."
        for document in documents[:8]
    ]
    ai_briefing = _read_only_briefing(
        document_signals,
        "Document posture updates once files and metadata enter the library.",
    )
    return schemas.DocumentsLibraryResponse(
        total_documents=db.query(models.Document).count(),
        documents=documents,
        document_signals=document_signals,
        ai_briefing=ai_briefing,
    )


@app.get("/operations/overview", response_model=schemas.OperationsOverviewResponse)
def operations_overview(
    db: Session = Depends(get_db), identity: tuple = Depends(require_auth)
) -> schemas.OperationsOverviewResponse:
    _username, _ = identity
    _ensure_workflow_records(db)
    tasks = [
        f"{task.title} [{task.priority}] owned by {task.owner_name or 'unassigned'}."
        for task in db.query(models.WorkflowTask).order_by(models.WorkflowTask.id.desc()).limit(12).all()
    ]
    exceptions = [
        f"{item.title} ({item.severity}) - {item.summary or 'review required'}."
        for item in db.query(models.WorkflowException).order_by(models.WorkflowException.id.desc()).limit(8).all()
    ]
    automation_priorities = [
        "Enforce stage gates before committee promotion.",
        "Generate owner digests from task and exception queues.",
        "Push escalation alerts to Slack or Teams once credentials are available.",
        "Route AI document reviews into the asynchronous processing lane.",
    ]
    ai_briefing = _read_only_briefing(
        exceptions + tasks + automation_priorities,
        "Workflow posture will appear once tasks and exceptions are active.",
    )
    return schemas.OperationsOverviewResponse(
        tasks=tasks,
        exceptions=exceptions,
        automation_priorities=automation_priorities,
        ai_briefing=ai_briefing,
    )


@app.get("/governance/overview", response_model=schemas.GovernanceOverviewResponse)
def governance_overview(
    db: Session = Depends(get_db), _: tuple = Depends(require_roles("admin", "manager"))
) -> schemas.GovernanceOverviewResponse:
    audit_events = [
        f"{event.action} on {event.entity_type} by {event.actor}: {event.detail or 'no detail'}"
        for event in db.query(models.AuditEvent).order_by(models.AuditEvent.id.desc()).limit(12).all()
    ]
    controls = [
        "Ollama is the default AI provider unless a governed enterprise override is configured.",
        "Microsoft Entra SSO can be activated through environment-backed tenant configuration.",
        "Integration toggles are stored server-side and reflected in the operations control plane.",
        "Signed session tokens expire automatically and can be replaced by stricter gateway enforcement later.",
    ]
    guardrails = [
        "Deal-stage changes and capital-touch actions require authenticated humans; AI outputs are advisory unless explicitly wired to automation with policy.",
        "Document Q&A is grounded on ingested deal text when chunks exist; empty retrieval yields a refusal-style answer from the model path.",
        "Copilot prompts are logged under AIRun for traceability; tune retention for your compliance regime.",
        "Integration toggles prevent placeholder vendors from impersonating live connectivity.",
    ]
    hallucination_controls = [
        "Prefer RAG + citations on deal-scoped questions; review citation list before acting.",
        "Use REOS_AI_MODE=local_fallback for deterministic offline demos without calling Ollama.",
        "Track override rate: when analysts ignore AI text, capture reason in deal notes for prompt and tool updates.",
        "Add gold-set regression tests before expanding prompt surface area (see backend/tests).",
    ]
    external_placeholders = [
        "Microsoft Graph / Gmail: inbox signals (not connected until OAuth env configured).",
        "Apollo.io: live prospect sync (requires REOS_APOLLO_API_KEY + worker).",
        "Azure OpenAI / AI Search: enterprise LLM and vector index (env-gated).",
        "Market data vendors (CompStak, etc.): underwriting comps (catalog only until contracted).",
        "Predictive forecasting models: train after stable labels from closed deals; not shipped in this repo yet.",
    ]
    ollama_line = (
        "Ollama reachable at configured URL."
        if current_ai_provider() == "ollama" and ollama_server_reachable()
        else (
            "Ollama not reachable or not selected; use local_fallback or start Ollama for generative paths."
            if current_ai_provider() == "ollama"
            else f"AI provider is {current_ai_provider()}; Ollama probe skipped."
        )
    )
    ai_runtime_notes = [
        ollama_line,
        f"Chat/generate model route: {current_ai_model()}.",
        "Structured financial verdicts (IRR, yield tables) belong in a dedicated tool layer; LLM explains, calculators compute.",
    ]
    return schemas.GovernanceOverviewResponse(
        audit_events=audit_events,
        ai_run_count=db.query(models.AIRun).count(),
        session_count=db.query(models.UserSession).count(),
        controls=controls,
        guardrails=guardrails,
        hallucination_controls=hallucination_controls,
        external_placeholders=external_placeholders,
        ai_runtime_notes=ai_runtime_notes,
    )


@app.get("/admin/overview", response_model=schemas.AdminOverviewResponse)
def admin_overview(db: Session = Depends(get_db), _: tuple = Depends(require_roles("admin"))) -> schemas.AdminOverviewResponse:
    status = _integration_status_snapshot()
    users = [
        f"{user.display_name or user.username} [{user.role}] via {user.provider}"
        for user in db.query(models.User).order_by(models.User.id.desc()).limit(15).all()
    ]
    providers = [
        "Microsoft Entra" if microsoft_auth_configured() else "Microsoft Entra (configuration pending)",
        "Local recovery access" if local_recovery_enabled() else "Local recovery access disabled",
    ]
    integration_status = [
        f"AI provider: {status.ai_provider}",
        f"Runtime mode: {status.runtime_mode}",
        f"Automation mode: {status.automation_mode}",
    ]
    operating_modes = [
        "Operator mode for command center decisions",
        "Governed mode for audit-heavy environments",
        "Hybrid local-to-enterprise mode for staged rollout",
    ]
    return schemas.AdminOverviewResponse(
        users=users,
        providers=providers,
        integration_status=integration_status,
        operating_modes=operating_modes,
    )


@app.get("/reports/overview", response_model=schemas.ReportsOverviewResponse)
def reports_overview(
    db: Session = Depends(get_db), identity: tuple = Depends(require_auth)
) -> schemas.ReportsOverviewResponse:
    _username, _ = identity
    nd = db.query(models.Deal).count()
    nic = db.query(models.DiligenceItem).filter(models.DiligenceItem.status != "done").count()
    nw = db.query(models.WorkflowException).filter(models.WorkflowException.status == "open").count()
    report_queue = [
        "Weekly IC briefing",
        "Investor outreach momentum report",
        "Document risk digest",
        "Workflow exception summary",
    ]
    if nd:
        report_queue.insert(
            0,
            f"Portfolio snapshot: {nd} deals, {nic} open diligence items, {nw} open workflow exceptions.",
        )
    executive_brief = _read_only_briefing(
        report_queue,
        "Executive reporting becomes available when portfolio data is present.",
    )
    return schemas.ReportsOverviewResponse(report_queue=report_queue, executive_brief=executive_brief)


@app.post("/ai/copilot", response_model=schemas.AICopilotResponse)
async def ai_copilot(
    payload: schemas.AICopilotRequest,
    db: Session = Depends(get_db),
    identity: tuple = Depends(require_auth),
) -> schemas.AICopilotResponse:
    username, _ = identity
    workspace = payload.workspace.strip().lower()
    if workspace == "deal" and payload.deal_id:
        deal = db.query(models.Deal).filter(models.Deal.id == payload.deal_id).first()
        if not deal:
            raise HTTPException(status_code=404, detail="Deal not found")
        context_text = "\n".join(
            [
                f"Deal {deal.name} is at stage {deal.stage} with priority {deal.priority}.",
                *(doc.summary or doc.filename for doc in db.query(models.Document).filter(models.Document.deal_id == deal.id).limit(8).all()),
                *(item.title for item in db.query(models.DiligenceItem).filter(models.DiligenceItem.deal_id == deal.id).limit(8).all()),
            ]
        )
    elif workspace == "deals":
        context_text = "\n".join(
            f"{deal.name} is at {deal.stage} with priority {deal.priority} and next action {deal.next_action or 'not set'}."
            for deal in db.query(models.Deal).order_by(models.Deal.id.desc()).limit(14).all()
        )
    elif workspace == "leads":
        lead_deals = (
            db.query(models.Deal)
            .filter(models.Deal.stage.in_(["Lead", "Screening"]))
            .order_by(models.Deal.id.desc())
            .limit(12)
            .all()
        )
        context_text = "\n".join(
            f"{deal.name} came from {deal.source or 'direct channel'} in {deal.city or 'unknown city'}."
            for deal in lead_deals
        )
    elif workspace == "crm":
        crm_contacts = (
            db.query(models.Contact)
            .options(joinedload(models.Contact.company))
            .order_by(models.Contact.id.desc())
            .limit(12)
            .all()
        )
        context_text = "\n".join(
            f"{contact.full_name} is a {contact.contact_type} linked to "
            f"{(contact.company.name if contact.company else 'direct relationship')}."
            for contact in crm_contacts
        )
    elif workspace == "documents":
        context_text = "\n".join(
            f"{document.filename} is {document.status} with {document.risk_tags or 'no'} risk tags."
            for document in db.query(models.Document).order_by(models.Document.id.desc()).limit(12).all()
        )
    elif workspace == "investors":
        context_text = "\n".join(
            f"{entry.status} with {entry.conviction or 'unknown'} conviction and next action {entry.next_action or 'not set'}."
            for entry in db.query(models.InvestorPipelineEntry).order_by(models.InvestorPipelineEntry.id.desc()).limit(12).all()
        )
    elif workspace == "operations":
        _ensure_workflow_records(db)
        context_text = "\n".join(
            task.title for task in db.query(models.WorkflowTask).order_by(models.WorkflowTask.id.desc()).limit(12).all()
        )
    else:
        context_text = _portfolio_context(db)
    answer = await _ai_briefing(db, workspace, username, context_text, payload.prompt)
    _record_audit_event(
        db,
        actor=username,
        action="copilot",
        entity_type=workspace,
        entity_id=str(payload.deal_id) if payload.deal_id else None,
        detail="Ran workspace copilot.",
    )
    db.commit()
    return schemas.AICopilotResponse(
        answer=answer,
        provider=current_ai_provider(),
        model_name=current_ai_model(),
        workspace=workspace,
    )
