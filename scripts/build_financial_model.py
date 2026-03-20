# Author: Victor.I
"""
Build reos_financial_model.xlsx: revenue, costs, cash, runway.
Edit ASSUMPTIONS below (or duplicate the workbook and replace numbers), then run:

  python scripts/build_financial_model.py

Output: docs/financial-model/reos_financial_model.xlsx
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_PATH = Path(__file__).resolve().parent.parent / "docs" / "financial-model" / "reos_financial_model.xlsx"

# --- Driver assumptions (replace with your plan; numbers are illustrative only) ---
ASSUMPTIONS = {
    "model_name": "REOS-style enterprise / proptech (illustrative)",
    "currency": "USD",
    "start_month_label": "2026-01",
    "starting_cash": 2_500_000,
    "initial_mrr_usd": 18_000,
    "mrr_growth_m1_m12": 0.075,
    "mrr_growth_m13_m24": 0.045,
    "mrr_growth_m25_m36": 0.025,
    "cogs_pct_of_revenue": 0.16,
    "base_payroll_monthly": 195_000,
    "payroll_step_add_every_6_months": 22_000,
    "infra_hosting_tools_monthly": 14_000,
    "ga_office_legal_ops_monthly": 42_000,
    "sm_pct_of_revenue": 0.22,
    "one_time_setup_month1": 95_000,
    "one_time_setup_label": "Legal, SOC2 prep, initial implementations (month 1)",
}


def mrr_for_month(month_index: int) -> float:
    """month_index 1..36"""
    mrr = ASSUMPTIONS["initial_mrr_usd"]
    for m in range(1, month_index + 1):
        if m <= 12:
            g = ASSUMPTIONS["mrr_growth_m1_m12"]
        elif m <= 24:
            g = ASSUMPTIONS["mrr_growth_m13_m24"]
        else:
            g = ASSUMPTIONS["mrr_growth_m25_m36"]
        mrr *= 1 + g
    return round(mrr, 2)


def payroll_for_month(month_index: int) -> float:
    steps = (month_index - 1) // 6
    return ASSUMPTIONS["base_payroll_monthly"] + steps * ASSUMPTIONS["payroll_step_add_every_6_months"]


def build() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Sheet: Assumptions (inputs used for this run) ---
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a["A1"] = "Financial model drivers"
    ws_a["A1"].font = Font(bold=True, size=14)
    ws_a["A3"] = "Parameter"
    ws_a["B3"] = "Value"
    ws_a["C3"] = "Notes"
    for col in ("A", "B", "C"):
        ws_a[f"{col}3"].font = Font(bold=True)
        ws_a[f"{col}3"].fill = PatternFill("solid", fgColor="DDDDDD")

    rows = [
        ("Starting cash (bank)", ASSUMPTIONS["starting_cash"], "Seed + early ARR; see assumptions doc"),
        ("Initial MRR", ASSUMPTIONS["initial_mrr_usd"], "Day-one contracted recurring revenue"),
        ("MRR growth months 1-12 (mo/mo)", ASSUMPTIONS["mrr_growth_m1_m12"], "New logos + expansion"),
        ("MRR growth months 13-24 (mo/mo)", ASSUMPTIONS["mrr_growth_m13_m24"], "Slower as base grows"),
        ("MRR growth months 25-36 (mo/mo)", ASSUMPTIONS["mrr_growth_m25_m36"], "Mature phase"),
        ("COGS % of revenue", ASSUMPTIONS["cogs_pct_of_revenue"], "Hosting, support load, pro services variable"),
        ("Base payroll (monthly)", ASSUMPTIONS["base_payroll_monthly"], "Loaded cost incl. benefits/taxes proxy"),
        ("Payroll add every 6 months", ASSUMPTIONS["payroll_step_add_every_6_months"], "Incremental hires / market adjustments"),
        ("Infra & tools (monthly)", ASSUMPTIONS["infra_hosting_tools_monthly"], "Cloud, observability, seats"),
        ("G&A (monthly)", ASSUMPTIONS["ga_office_legal_ops_monthly"], "Finance, HR, insurance, misc"),
        ("S&M % of revenue", ASSUMPTIONS["sm_pct_of_revenue"], "Sales + marketing as % of recognized rev"),
        ("One-time month-1 cash", ASSUMPTIONS["one_time_setup_month1"], ASSUMPTIONS["one_time_setup_label"]),
    ]
    for i, (p, v, n) in enumerate(rows, start=4):
        ws_a.cell(row=i, column=1, value=p)
        ws_a.cell(row=i, column=2, value=v)
        ws_a.cell(row=i, column=3, value=n)
    ws_a.column_dimensions["A"].width = 38
    ws_a.column_dimensions["B"].width = 22
    ws_a.column_dimensions["C"].width = 55

    # --- Sheet: Monthly projections ---
    ws_m = wb.create_sheet("Monthly_P_L")
    headers = [
        "Month #",
        "MRR",
        "Revenue (accrued = MRR)",
        "COGS",
        "Gross profit",
        "Payroll",
        "S&M",
        "Infra",
        "G&A",
        "One-time / other",
        "Total opex",
        "Net operating cash (before WC)",
        "Beginning cash",
        "Ending cash",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws_m.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDDDDD")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    cash = ASSUMPTIONS["starting_cash"]
    runway_hit = None

    for m in range(1, 37):
        row = m + 1
        rev = mrr_for_month(m)
        cogs = round(rev * ASSUMPTIONS["cogs_pct_of_revenue"], 2)
        gross = round(rev - cogs, 2)
        payroll = payroll_for_month(m)
        sm = round(rev * ASSUMPTIONS["sm_pct_of_revenue"], 2)
        infra = ASSUMPTIONS["infra_hosting_tools_monthly"]
        ga = ASSUMPTIONS["ga_office_legal_ops_monthly"]
        one_time = ASSUMPTIONS["one_time_setup_month1"] if m == 1 else 0
        opex = round(payroll + sm + infra + ga + one_time, 2)
        net = round(gross - opex, 2)
        beg = cash
        cash = round(beg + net, 2)
        if cash <= 0 and runway_hit is None:
            runway_hit = m - 1

        ws_m.cell(row=row, column=1, value=m)
        ws_m.cell(row=row, column=2, value=rev)
        ws_m.cell(row=row, column=3, value=rev)
        ws_m.cell(row=row, column=4, value=cogs)
        ws_m.cell(row=row, column=5, value=gross)
        ws_m.cell(row=row, column=6, value=payroll)
        ws_m.cell(row=row, column=7, value=sm)
        ws_m.cell(row=row, column=8, value=infra)
        ws_m.cell(row=row, column=9, value=ga)
        ws_m.cell(row=row, column=10, value=one_time)
        ws_m.cell(row=row, column=11, value=opex)
        ws_m.cell(row=row, column=12, value=net)
        ws_m.cell(row=row, column=13, value=beg)
        ws_m.cell(row=row, column=14, value=cash)

    for col in range(1, 15):
        ws_m.column_dimensions[get_column_letter(col)].width = 16

    # --- Sheet: Annual rollup ---
    ws_y = wb.create_sheet("Annual_summary")
    ws_y["A1"] = "Year"
    ws_y["B1"] = "Revenue (sum MRR-months as proxy)"
    ws_y["C1"] = "Total opex (excl. WC)"
    ws_y["D1"] = "Net cash flow (simple)"
    for col in ("A", "B", "C", "D"):
        ws_y[f"{col}1"].font = Font(bold=True)

    def sum_months(start_m: int, end_m: int) -> tuple[float, float, float]:
        tr = to = tnet = 0.0
        for m in range(start_m, end_m + 1):
            rev = mrr_for_month(m)
            cogs = rev * ASSUMPTIONS["cogs_pct_of_revenue"]
            gross = rev - cogs
            payroll = payroll_for_month(m)
            sm = rev * ASSUMPTIONS["sm_pct_of_revenue"]
            infra = ASSUMPTIONS["infra_hosting_tools_monthly"]
            ga = ASSUMPTIONS["ga_office_legal_ops_monthly"]
            one_time = ASSUMPTIONS["one_time_setup_month1"] if m == 1 else 0
            opex = payroll + sm + infra + ga + one_time
            tr += rev
            to += opex
            tnet += gross - opex
        return round(tr, 2), round(to, 2), round(tnet, 2)

    y1r, y1o, y1n = sum_months(1, 12)
    y2r, y2o, y2n = sum_months(13, 24)
    y3r, y3o, y3n = sum_months(25, 36)
    for i, (yl, r, o, n) in enumerate(
        [(1, y1r, y1o, y1n), (2, y2r, y2o, y2n), (3, y3r, y3o, y3n)], start=2
    ):
        ws_y.cell(row=i, column=1, value=f"Year {yl}")
        ws_y.cell(row=i, column=2, value=r)
        ws_y.cell(row=i, column=3, value=o)
        ws_y.cell(row=i, column=4, value=n)

    # --- Sheet: Runway ---
    ws_r = wb.create_sheet("Runway")
    ws_r["A1"] = "Runway analysis (simplified)"
    ws_r["A1"].font = Font(bold=True, size=12)
    ws_r["A3"] = "Starting cash"
    ws_r["B3"] = ASSUMPTIONS["starting_cash"]
    ws_r["A4"] = "Cash after month 36"
    ws_r["B4"] = f"=Monthly_P_L!N38"
    ws_r["A5"] = "First month ending cash <= 0 (if any)"
    ws_r["B5"] = runway_hit if runway_hit else "Not hit in 36 mo (under these drivers)"
    ws_r["A7"] = "Notes"
    ws_r["B7"] = (
        "No AR/AP, deferred revenue, or debt service modeled. Net line is crude operating cash proxy. "
        "Extend Monthly_P_L or add a Cash_Flow sheet for fundraise timing."
    )
    ws_r.column_dimensions["A"].width = 40
    ws_r.column_dimensions["B"].width = 50

    # Fix B4 to value because row 38 = month 36 ending - compute in python for reliability
    cash_end = cash
    ws_r["B4"] = cash_end
    ws_r["B4"].number_format = "#,##0"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    build()
